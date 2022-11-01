from openpyxl.worksheet.worksheet import Worksheet

from Nucarnival.cardHelper import CardHelper
from Nucarnival.nucarnivalHelper import NucarnivalHelper
from RoleCards.cards.blade.explosiveRecall import ExplosiveRecall
from RoleCards.cards.edmond.knightlyNight import KnightlyNight
from RoleCards.cards.edmond.sweetAroma import SweetAroma
from RoleCards.cards.garu.howlingCyclone import HowlingCyclone
from RoleCards.cards.kuya.lakesideSpark import LakesideSpark
from RoleCards.cards.monster.commonMonster import CommonMonster
from RoleCards.cards.monster.tempTeamMate import TempTeamMate
from RoleCards.cards.olivine.aquaBloom import AquaBloom
from RoleCards.cards.olivine.manOfGod import ManOfGod
from RoleCards.cards.olivine.radiantAdmiral import RadiantAdmiral
from RoleCards.cards.quincy.ancientCeremony import AncientCeremony
from RoleCards.cards.quincy.distantPromise import DistantPromise
from RoleCards.cards.yakumo.homecoming import Homecoming
from RoleCards.cards.yakumo.oceanBreeze import OceanBreeze
from RoleCards.common.card import ICard
from RoleCards.enum.cardOccupationEnum import CardOccupation
from RoleCards.enum.cardRoleEnum import CardRole
from RoleCards.enum.cardTypeEnum import CardType
from openpyxl.workbook import Workbook

from tkinter import filedialog


# 模拟角色单人13回合作战能力
def simulation1(cardHelper: CardHelper, helper: NucarnivalHelper):
    wb = Workbook()

    ws = wb.create_sheet('计算结果', 0)

    ws.cell(1, 1, '卡')
    ws.cell(1, 2, '昵称')
    ws.cell(1, 3, '角色')
    ws.cell(1, 4, '属性')
    ws.cell(1, 5, '定位')
    ws.cell(1, 6, '星级')
    ws.cell(1, 7, '潜能')
    ws.cell(1, 8, 'Hp')
    ws.cell(1, 9, 'Atk')
    ws.cell(1, 10, '13回合单人输出')

    row = 1
    for x in cardHelper.cardList:
        if (x.occupation == CardOccupation.Support and x.cardName != '诡夜疾风') or x.occupation == CardOccupation.Healer:
            continue
        if x.isGroup is False:
            row += 1
            # 5星满潜
            simulation(helper, ws, x, row, 5, 12)

            row += 1
            # 5星6潜
            simulation(helper, ws, x, row, 5, 6)

            row += 1
            # 3星满潜
            simulation(helper, ws, x, row, 3, 12)

            row += 1
            # 3星6潜
            simulation(helper, ws, x, row, 3, 6)


    wb.save('C:\\fhs\\python\\计算结果2.xls')


# 进行模拟
def simulation(helper: NucarnivalHelper, ws: Worksheet, x: ICard, row, _star, _tier):
    x.setProperties(60, _star, 5, _tier)
    x.calHpAtk()
    export(ws, x, row)

    helper.clearUp()
    similationTeamMate(helper, x)
    # srO = ManOfGod()
    # srO.setProperties(60, 5, 5, 12)
    # srO.calHpAtk()
    # helper.team.append(srO)
    helper.team.append(x)
    helper.monsters.append(CommonMonster())
    helper.maxTurn = 13
    helper.battleStart(False)
    ws.cell(row, 10, helper.damageRecord[x])


# 导出
def export(ws: Worksheet, x: ICard, row):
    ws.cell(row, 1, x.cardName)
    ws.cell(row, 2, x.nickName)
    ws.cell(row, 3, x.role.value)
    ws.cell(row, 4, x.type.typeName)
    ws.cell(row, 5, x.occupation.occupationName)
    ws.cell(row, 6, x.star)
    ws.cell(row, 7, x.tier)
    ws.cell(row, 8, x.hp)
    ws.cell(row, 9, x.atk)


# 配置模拟队友
def similationTeamMate(helper: NucarnivalHelper, x: ICard):
    if x.cardName == '爵士册封之夜' or x.cardName == '追逐悠远之约':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Striker
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.occupation = CardOccupation.Striker
        helper.team.append(mate2)
    elif x.cardName == '化形宴，飘落叶' or x.cardName == '诡秘妖狐':
        mate = TempTeamMate()
        mate.role = CardRole.Kuya
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Kuya
        helper.team.append(mate2)
    elif x.cardName == '来自主人的新礼物' or x.cardName == '流浪小狼':
        mate = TempTeamMate()
        mate.role = CardRole.Garu
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Garu
        helper.team.append(mate2)
    elif x.cardName == '战斗人偶与回忆' or x.cardName == '魔人偶':
        mate = TempTeamMate()
        mate.role = CardRole.Blade
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Blade
        helper.team.append(mate2)
    elif x.cardName == '白色恋人':
        mate = TempTeamMate()
        mate.role = CardRole.Yakumo
        helper.team.append(mate)
    elif x.cardName == '七叶之花的奇迹':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Support
        helper.team.append(mate)
    elif x.cardName == '华宴夜未央':
        mate = TempTeamMate()
        mate.role = CardRole.Kuya
        helper.team.append(mate)
    elif x.cardName == '偶像实习':
        mate = TempTeamMate()
        mate.role = CardRole.Yakumo
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Olivine
        helper.team.append(mate2)
    elif x.cardName == '永不消散的焰火' or x.cardName == '太阳城主':
        mate = TempTeamMate()
        mate.role = CardRole.Dante
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Dante
        helper.team.append(mate2)
    elif x.cardName == '初露芬芳的甜味':
        mate = TempTeamMate()
        helper.team.append(mate)
    elif x.cardName == '使魔-艾斯特':
        mate = TempTeamMate()
        mate.role = CardRole.Morvay
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Aster
        helper.team.append(mate2)
        mate3 = TempTeamMate()
        mate3.role = CardRole.Aster
        helper.team.append(mate3)
    elif x.cardName == '骑士副团长':
        mate = TempTeamMate()
        mate.role = CardRole.Edmond
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Edmond
        helper.team.append(mate2)
    elif x.cardName == '古森守护者':
        mate = TempTeamMate()
        mate.role = CardRole.Quincy
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Quincy
        helper.team.append(mate2)


if __name__ == '__main__':
    _helper = NucarnivalHelper()

    _cardHelper = CardHelper()
    # simulation1(_cardHelper, _helper)

    yk = DistantPromise()
    yk.setProperties(60, 3, 5, 12)
    yk.calHpAtk()

    pK = AncientCeremony()
    pK.setProperties(60, 3, 5, 12)
    pK.calHpAtk()

    #
    _helper.clearUp()
    # _helper.team.append(hO)
    # _helper.team.append(xO)
    _helper.team.append(pK)
    # _helper.team.append(pB)
    similationTeamMate(_helper, pK)
    _helper.monsters.append(CommonMonster())
    _helper.maxTurn = 13

    _helper.battleStart(True)
    # _helper.exportExcel('计算结果','E:\\新世界\\计算结果.xls')
