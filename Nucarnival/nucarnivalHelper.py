import io
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from Common.nvEventManager import EventListener, Event, EventType, eventManagerInstance
from RoleCards.common.card import ICard, writeCardInfoTitleInExcel
from RoleCards.enum.buffTypeEnum import BuffType
from RoleCards.enum.conditionTypeEnum import ConditionType
from Common.ncRound import roundHalfEven, roundDown


class NucarnivalHelper:
    def __init__(self):
        self.team: list[ICard] = []
        self.monsters: list[ICard] = []
        self.maxTurn = 50
        self.defenseTurn = {}
        self.skillTurn = {}
        self.currentTurn = 0
        self.totalDamage = 0
        self.output = io.StringIO()
        self.wb: Workbook = None
        self.ws: Worksheet = None
        self.roleWS: Worksheet = None
        self.columnMark = {}
        self.battleListener = BattleRecordListener()
        eventManagerInstance.addListener(self.battleListener)

    def clearUp(self):
        self.team: list[ICard] = []
        self.monsters: list[ICard] = []
        self.maxTurn = 50
        self.defenseTurn = {}
        self.skillTurn = {}
        self.clearUpBattleResult()

    def clearUpBattleResult(self):
        self.battleListener.cleanUp()
        self.currentTurn = 0
        self.totalDamage = 0
        self.output.close()
        self.output = io.StringIO()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '伤害模拟结果'
        self.roleWS = self.wb.create_sheet('角色属性', 1)

    def recordBattleMsg(self, msg: str):
        self.output.seek(0, 2)
        self.output.write(msg + '\n')

    # 战斗开始
    # 双方清除旧buff、添加队友和敌方信息
    # 激活双方buff
    # 我方行动：防御/普攻/必杀 -> 结算dot ->结算hot
    # 敌方行动：防御/普攻/必杀 -> 结算dot ->结算hot
    # 我方全部阵亡/敌方全部阵亡-> 退出战斗
    def battleStart(self, printInfo=False):
        self.clearUpBattleResult()

        self.ws.merge_cells(None, 1, 1, 3, 1)
        self.ws.cell(1, 1, '回合')

        comment = Comment('为了简略Excel，部分行动将用缩写'
                          '\n持续伤害=dot'
                          '\n持续治疗=hot', '纳萨尔')
        self.ws.cell(1, 1).comment = comment

        column = 2
        for role in self.team:
            role.teamMate = self.team
            role.enemies = self.monsters
            role.clearUp()
            role.calHpAtk()
            self.columnMark[role] = column
            self.ws.merge_cells(None, 1, column, 1, column + 2)
            self.ws.cell(1, column, role.exportCardInfo())
            self.ws.cell(1, column).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

            self.ws.merge_cells(None, 2, column, 3, column)
            self.ws.cell(2, column, '实时攻击力')
            self.ws.cell(2, column).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

            self.ws.merge_cells(None, 2, column + 1, 2, column + 2)
            self.ws.cell(2, column + 1, '回合行动')
            self.ws.cell(3, column + 1, '伤害')
            self.ws.cell(3, column + 2, '治疗')
            column += 3
        for monster in self.monsters:
            monster.teamMate = self.monsters
            monster.enemies = self.team
            monster.clearUp()
        self.writeCardInfoInExcel()

        if self.maxTurn > 50:
            self.maxTurn = 50
        elif self.maxTurn < 1:
            self.maxTurn = 1

        msg = '-----------开始------------'
        self.recordBattleMsg(msg)

        for turn in range(0, self.maxTurn + 1):
            self.currentTurn = turn
            self.battleListener.currentTurn = turn
            if turn == 0:
                for role in self.team:
                    role.activeBuffs()
                for tempMonster in self.monsters:
                    tempMonster.activeBuffs()
                for role in self.team:
                    role.calMaxHp()
                for tempMonster in self.monsters:
                    tempMonster.calMaxHp()
                continue

            if turn > 1:
                self.recordBattleMsg('')
            msg = '---------第{0}回合---------'.format(turn)
            row = 3
            self.ws.cell(row + turn, 1, '{0}'.format(turn))
            self.recordBattleMsg(msg)

            # 我方行动
            self.action(turn, self.team, self.monsters)

            for role in self.team:
                role.settleDot()
                role.settleHot()

            # 敌方行动
            self.action(turn, self.monsters, self.team)

            for role in self.monsters:
                role.settleDot()
                role.settleHot()

            # 下个回合
            temp = 0
            for role in self.team:
                if temp != 0:
                    self.recordBattleMsg('')
                self.recordResult(role, turn, row)
                role.nextRound()
                temp += 1
            for monster in self.monsters:
                monster.nextRound()

        msg = '-----------结束------------'
        self.recordBattleMsg(msg)
        self.totalDamage = 0
        self.ws.cell(row + turn + 1, 1, '统计')
        self.ws.cell(row + turn + 2, 1, '占比')
        self.ws.cell(row + turn + 3, 1, '总{}回合\n全队总伤害'.format(turn))
        self.ws.cell(row + turn + 3, 1,).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        self.recordBattleMsg('总{}回合'.format(turn))
        totalDamage = 0
        temp = 0
        maxColumn = 0
        for role in self.team:
            if temp != 0:
                self.recordBattleMsg('')
            temp += 1
            totalDamage += self.recordTotalResult(role, turn, row)
            if maxColumn < self.columnMark[role]:
                maxColumn = self.columnMark[role]
        msg = '全队伤害：{}'.format(totalDamage)
        self.ws.merge_cells(None, row + turn + 3, 2, row + turn + 3, maxColumn + 2)
        self.ws.cell(row + turn + 3, 2, totalDamage)
        self.ws.cell(row + turn + 3, 2).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        self.recordBattleMsg(msg)
        if printInfo:
            print(self.output.getvalue())

    # 获取本回合总伤害
    def recordResult(self, role: ICard, turn, row):
        msg = role.cardInfo(False)
        wsMsgD = ''
        wsMsgH = ''
        isDefense = False

        turnAtk = role.getCurrentAtk()
        eventAtk = Event(EventType.turnAtk)
        eventAtk.data['source'] = role
        eventAtk.data['value'] = turnAtk
        eventAtk.data['target'] = role
        eventManagerInstance.sendEvent(eventAtk)

        atkMag = '行动前：'
        atkAfterMag = '行动后：'
        atkTurnMag = '回合：{}'.format(turnAtk)

        if len(self.battleListener.findSourceEvent(role, EventType.actionAtk)) > 0:
            try:
                actionAtk = self.battleListener.findSourceEvent(role, EventType.actionAtk)[0].value
                msg += '[{}]'.format(actionAtk)
                atkMag += str(actionAtk)
            except:
                print('获取行动实时攻击力记录出错')

        if len(self.battleListener.findSourceEvent(role, EventType.defense)) > 0:
            msg += '  防御'
            if len(wsMsgD) > 0:
                wsMsgD += '\n'
            wsMsgD = '防御'
            if len(wsMsgH) > 0:
                wsMsgH += '\n'
            wsMsgH = '防御'
            isDefense = True

        attackDamage = 0
        for record in self.battleListener.findSourceEvent(role, EventType.attackDamage):
            try:
                attackDamage += record.value
            except:
                print('获取普攻记录出错')
        attackFU = 0
        for record in self.battleListener.findSourceEvent(role, EventType.attackFollowUp):
            try:
                attackFU += record.value
            except:
                print('获取普攻追击记录出错')
        attackHeals = {}
        for healRole in self.team:
            attackHeals[healRole] = 0
        attackHeal = 0
        for record in self.battleListener.findSourceEvent(role, EventType.attackHeal):
            try:
                attackHeal += record.value
                if record.target is not None and record.target in self.team:
                    attackHeal2 = attackHeals[record.target]
                    attackHeal2 += record.value
                    attackHeals[record.target] = attackHeal2
            except:
                print('获取普攻治疗记录出错')
        skillDamage = 0
        for record in self.battleListener.findSourceEvent(role, EventType.skillDamage):
            try:
                skillDamage += record.value
            except:
                print('获取必杀记录出错')
        skillFU = 0
        for record in self.battleListener.findSourceEvent(role, EventType.skillFollowUp):
            try:
                skillFU += record.value
            except:
                print('获取必杀追击记录出错')
        skillHeals = {}
        for healRole in self.team:
            skillHeals[healRole] = 0
        skillHeal = 0
        for record in self.battleListener.findSourceEvent(role, EventType.skillHeal):
            try:
                skillHeal += record.value
                if record.target is not None and record.target in self.team:
                    skillHeal2 = skillHeals[record.target]
                    skillHeal2 += record.value
                    skillHeals[record.target] = skillHeal2
            except:
                print('获取必杀治疗记录出错')
        dot = 0
        for record in self.battleListener.findSourceEvent(role, EventType.dot):
            try:
                dot += record.value
            except:
                print('获取持续伤害记录出错')
        hotHeals = {}
        for healRole in self.team:
            hotHeals[healRole] = 0
        hot = 0
        for record in self.battleListener.findSourceEvent(role, EventType.hot):
            try:
                hot += record.value
                if record.target is not None and record.target in self.team:
                    hotHeal = hotHeals[record.target]
                    hotHeal += record.value
                    hotHeals[record.target] = hotHeal
            except:
                print('获取持续治疗记录出错')
        bloodSuck = 0
        for record in self.battleListener.findSourceEvent(role, EventType.bloodSucking):
            try:
                bloodSuck += record.value
            except:
                print('获取吸血记录出错')
        shields = {}
        for healRole in self.team:
            shields[healRole] = 0
        shield = 0
        for record in self.battleListener.findSourceEvent(role, EventType.shield):
            try:
                shield += record.value
                if record.target is not None and record.target in self.team:
                    shield2 = shields[record.target]
                    shield2 += record.value
                    shields[record.target] = shield2
            except:
                print('获取护盾记录出错')
        counter = 0
        for record in self.battleListener.findSourceEvent(role, EventType.counter):
            try:
                counter += record.value
            except:
                print('获取反击记录出错')

        turnDamage = attackDamage + attackFU + skillDamage + skillFU + dot + counter
        turnHeal = attackHeal + skillHeal + hot
        if attackDamage > 0:
            astr = str(attackDamage)
            if attackFU > 0:
                astr += '(' + str(attackFU) + ')'
            msg += '  普攻造成伤害：' + astr
            if len(wsMsgD) > 0:
                wsMsgD += '\n'
            wsMsgD += '普攻：' + astr
        else:
            if attackFU > 0:
                msg += '  普攻造成伤害：' + str(attackFU)
                if len(wsMsgD) > 0:
                    wsMsgD += '\n'
                wsMsgD += '普攻：' + str(attackFU)
        if skillDamage > 0:
            astr = str(skillDamage)
            if skillFU > 0:
                astr += '(' + str(skillFU) + ')'
            msg += '  必杀造成伤害：' + astr
            if len(wsMsgD) > 0:
                wsMsgD += '\n'
            wsMsgD += '必杀：' + astr
        else:
            if skillFU > 0:
                msg += '  必杀造成伤害：' + str(skillFU)
                if len(wsMsgD) > 0:
                    wsMsgD += '\n'
                wsMsgD += '必杀：' + str(skillFU)
        if counter > 0:
            msg += '  反击造成伤害：' + str(counter)
            if len(wsMsgD) > 0:
                wsMsgD += '\n'
            wsMsgD += '反击：' + str(counter)
        if attackHeal > 0:
            healRoleCount = 0
            attackHealsMsg = ' ('
            for healRole in self.team:
                attackHeal2 = attackHeals[healRole]
                if attackHeal2 > 0:
                    if healRoleCount > 0:
                        attackHealsMsg += ','
                    healRoleCount += 1
                    attackHealsMsg += str(attackHeal2)
            attackHealsMsg += ') '
            msg += '  普攻造成治疗：' + str(attackHeal)
            if healRoleCount > 1:
                msg += attackHealsMsg
            if len(wsMsgH) > 0:
                wsMsgH += '\n'
            wsMsgH += '普攻：' + str(attackHeal)
        if skillHeal > 0:
            healRoleCount = 0
            skillHealsMsg = ' ('
            for healRole in self.team:
                skillHeal2 = skillHeals[healRole]
                if skillHeal2 > 0:
                    if healRoleCount > 0:
                        skillHealsMsg += ','
                    healRoleCount += 1
                    skillHealsMsg += str(skillHeal2)
            skillHealsMsg += ') '
            msg += '  必杀造成治疗：' + str(skillHeal)
            if healRoleCount > 1:
                msg += skillHealsMsg
            if len(wsMsgH) > 0:
                wsMsgH += '\n'
            wsMsgH += '必杀：' + str(skillHeal)
        if bloodSuck > 0:
            msg += '  吸血造成治疗：' + str(bloodSuck)
            if len(wsMsgH) > 0:
                wsMsgH += '\n'
            wsMsgH += '吸血：' + str(bloodSuck)
        if dot > 0:
            msg += '  造成持续伤害：' + str(dot)
            if len(wsMsgD) > 0:
                wsMsgD += '\n'
            wsMsgD += 'dot：' + str(dot)
        if hot > 0:
            healRoleCount = 0
            hotHealsMsg = ' ('
            for healRole in self.team:
                hotHeal = hotHeals[healRole]
                if hotHeal > 0:
                    if healRoleCount > 0:
                        hotHealsMsg += ','
                    healRoleCount += 1
                    hotHealsMsg += str(hotHeal)
            hotHealsMsg += ') '
            msg += '  造成持续治疗：' + str(hot)
            if healRoleCount > 1:
                msg += hotHealsMsg
            if len(wsMsgH) > 0:
                wsMsgH += '\n'
            wsMsgH += 'hot：' + str(hot)
        if shield > 0:
            healRoleCount = 0
            shieldMsg = ' ('
            for healRole in self.team:
                shield2 = shields[healRole]
                if shield2 > 0:
                    if healRoleCount > 0:
                        shieldMsg += ','
                    healRoleCount += 1
                    shieldMsg += str(shield2)
            shieldMsg += ') '
            msg += '  造成护盾：' + str(shield)
            if healRoleCount > 1:
                msg += shieldMsg
            if len(wsMsgH) > 0:
                wsMsgH += '\n'
            wsMsgH += '护盾：' + str(shield)
        if turnDamage > 0:
            if len(wsMsgD) > 0:
                wsMsgD += '\n'
            wsMsgD += '总伤：' + str(turnDamage)
        if turnHeal > 0:
            if len(wsMsgH) > 0:
                wsMsgH += '\n'
            wsMsgH += '总治：' + str(turnHeal)

        if turnDamage <= 0 and turnHeal <= 0 and isDefense is False:
            if len(self.battleListener.findSourceEvent(role, EventType.attack)) > 0:
                msg += '  普攻'
                if len(wsMsgD) > 0:
                    wsMsgD += '\n'
                wsMsgD = '普攻'
                if len(wsMsgH) > 0:
                    wsMsgH += '\n'
                wsMsgH = '普攻'
            if len(self.battleListener.findSourceEvent(role, EventType.skill)) > 0:
                msg += '  必杀'
                if len(wsMsgD) > 0:
                    wsMsgD += '\n'
                wsMsgD = '必杀'
                if len(wsMsgH) > 0:
                    wsMsgH += '\n'
                wsMsgH = '必杀'

        self.recordBattleMsg(msg)

        if len(self.battleListener.findSourceEvent(role, EventType.actionAfterAtk)) > 0:
            try:
                actionAtk = self.battleListener.findSourceEvent(role, EventType.actionAfterAtk)[0].value
                msg_atkAfter = '行动后实时攻击力[{}]'.format(actionAtk)
                self.recordBattleMsg(msg_atkAfter)
                atkAfterMag += str(actionAtk)
            except:
                print('获取行动后实时攻击力记录出错')
        msg_turnAtk = '我方所有角色行动后，本回合实时攻击力[{}]'.format(turnAtk)
        self.recordBattleMsg(msg_turnAtk)

        column = self.columnMark[role]
        msg3 = atkMag + '\n' + atkAfterMag + '\n' + atkTurnMag
        self.ws.cell(row + turn, column, msg3)
        self.ws.cell(row + turn, column).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        self.ws.cell(row + turn, column + 1, wsMsgD)
        self.ws.cell(row + turn, column + 1).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        self.ws.cell(row + turn, column + 2, wsMsgH)
        self.ws.cell(row + turn, column + 2).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    def getTotalResult(self, role: ICard):
        attackDamage = 0
        for record in self.battleListener.findSourceEvent(role, EventType.attackDamage, True):
            try:
                attackDamage += record.value
            except:
                print('获取普攻记录出错')
        attackFU = 0
        for record in self.battleListener.findSourceEvent(role, EventType.attackFollowUp, True):
            try:
                attackFU += record.value
            except:
                print('获取普攻追击记录出错')
        attackHeal = 0
        for record in self.battleListener.findSourceEvent(role, EventType.attackHeal, True):
            try:
                attackHeal += record.value
            except:
                print('获取普攻治疗记录出错')
        skillDamage = 0
        for record in self.battleListener.findSourceEvent(role, EventType.skillDamage, True):
            try:
                skillDamage += record.value
            except:
                print('获取必杀记录出错')
        skillFU = 0
        for record in self.battleListener.findSourceEvent(role, EventType.skillFollowUp, True):
            try:
                skillFU += record.value
            except:
                print('获取必杀追击记录出错')
        skillHeal = 0
        for record in self.battleListener.findSourceEvent(role, EventType.skillHeal, True):
            try:
                skillHeal += record.value
            except:
                print('获取必杀治疗记录出错')
        dot = 0
        for record in self.battleListener.findSourceEvent(role, EventType.dot, True):
            try:
                dot += record.value
            except:
                print('获取持续伤害记录出错')
        hot = 0
        for record in self.battleListener.findSourceEvent(role, EventType.hot, True):
            try:
                hot += record.value
            except:
                print('获取持续治疗记录出错')
        bloodSuck = 0
        for record in self.battleListener.findCustomEvent(role, EventType.bloodSucking, True):
            try:
                bloodSuck += record.value
            except:
                print('获取吸血记录出错')
        shield = 0
        for record in self.battleListener.findSourceEvent(role, EventType.shield, True):
            try:
                shield += record.value
            except:
                print('获取护盾记录出错')
        counter = 0
        for record in self.battleListener.findSourceEvent(role, EventType.counter, True):
            try:
                counter += record.value
            except:
                print('获取反击记录出错')
        totalDamage = attackDamage + attackFU + skillDamage + skillFU + dot + counter
        totalHeal = attackHeal + skillHeal + hot + bloodSuck
        data = {}
        data['attackDamage'] = attackDamage
        data['attackFU'] = attackFU
        data['skillDamage'] = skillDamage
        data['skillFU'] = skillFU
        data['dot'] = dot
        data['counter'] = counter
        data['attackHeal'] = attackHeal
        data['skillHeal'] = skillHeal
        data['hot'] = hot
        data['bloodSuck'] = bloodSuck
        data['shield'] = shield
        data['totalDamage'] = totalDamage
        data['totalHeal'] = totalHeal
        return data

    def recordTotalResult(self, role: ICard, turn, row):
        msg = role.cardInfo(False)
        data = self.getTotalResult(role)
        attackDamage = data['attackDamage']
        attackFU = data['attackFU']
        skillDamage = data['skillDamage']
        skillFU = data['skillFU']
        dot = data['dot']
        counter = data['counter']
        attackHeal = data['attackHeal']
        skillHeal = data['skillHeal']
        hot = data['hot']
        bloodSuck = data['bloodSuck']
        shield = data['shield']
        totalDamage = data['totalDamage']
        totalHeal = data['totalHeal']

        column = self.columnMark[role]
        self.ws.cell(row + turn + 1, column + 1, totalDamage)
        self.ws.cell(row + turn + 1, column + 1).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        self.ws.cell(row + turn + 1, column + 2, totalHeal)
        self.ws.cell(row + turn + 1, column + 2).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        if totalDamage > 0:
            msg += '  总伤害：{}'.format(totalDamage)
        if totalHeal > 0:
            msg += '  总治疗：{}'.format(totalHeal)
        if shield > 0:
            msg += '  总护盾：{}'.format(shield)
        if totalDamage > 0 or totalHeal > 0 or shield > 0:
            self.recordBattleMsg(msg)

        msg2 = '伤害占比'
        shzb = ''
        if totalDamage > 0:
            attack = attackDamage + attackFU
            skill = skillDamage + skillFU
            if attack > 0:
                proportion = roundHalfEven(attack / totalDamage * 100)
                if len(shzb) > 0:
                    shzb += '\n'
                shzb += '普攻({}%)'.format(proportion)
                msg2 += '  普攻({}%)'.format(proportion)
            if skill > 0:
                proportion = roundHalfEven(skill / totalDamage * 100)
                if len(shzb) > 0:
                    shzb += '\n'
                shzb += '必杀({}%)'.format(proportion)
                msg2 += '  必杀({}%)'.format(proportion)
            if dot > 0:
                proportion = roundHalfEven(dot / totalDamage * 100)
                if len(shzb) > 0:
                    shzb += '\n'
                shzb += 'dot({}%)'.format(proportion)
                msg2 += '  持续伤害({}%)'.format(proportion)
            if counter > 0:
                proportion = roundHalfEven(counter / totalDamage * 100)
                if len(shzb) > 0:
                    shzb += '\n'
                shzb += '反击({}%)'.format(proportion)
                msg2 += '  反击({}%)'.format(proportion)
            self.recordBattleMsg(msg2)
        self.ws.cell(row + turn + 2, column + 1, str(shzb))
        self.ws.cell(row + turn + 2, column + 1).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        zlzb = ''
        msg2 = '治疗占比'
        if totalHeal > 0:
            if attackHeal > 0:
                proportion = roundHalfEven(attackHeal / totalHeal * 100)
                if len(zlzb) > 0:
                    zlzb += '\n'
                zlzb += '普攻({}%)'.format(proportion)
                msg2 += '  普攻({}%)'.format(proportion)
            if skillHeal > 0:
                proportion = roundHalfEven(skillHeal / totalHeal * 100)
                if len(zlzb) > 0:
                    zlzb += '\n'
                zlzb += '必杀({}%)'.format(proportion)
                msg2 += '  必杀({}%)'.format(proportion)
            if hot > 0:
                proportion = roundHalfEven(hot / totalHeal * 100)
                if len(zlzb) > 0:
                    zlzb += '\n'
                zlzb += 'hot({}%)'.format(proportion)
                msg2 += '  持续治疗({}%)'.format(proportion)
            if bloodSuck > 0:
                proportion = roundHalfEven(bloodSuck / totalHeal * 100)
                if len(zlzb) > 0:
                    zlzb += '\n'
                zlzb += '吸血({}%)'.format(proportion)
                msg2 += '  吸血({}%)'.format(proportion)
            self.recordBattleMsg(msg2)
        self.ws.cell(row + turn + 2, column + 2, str(zlzb))
        self.ws.cell(row + turn + 2, column + 2).alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        return totalDamage

    # 行动
    # 防御/普攻/必杀
    # 结算dot
    # 结算hot
    def action(self, turn=0, cardList: list[ICard] = [], cardList2: list[ICard] = []):
        for role in cardList:
            if role in self.defenseTurn and turn in self.defenseTurn[role]:
                role.defense = True
                event = Event(EventType.defense)
                event.data['source'] = role
                event.data['value'] = 0
                event.data['target'] = role
                eventManagerInstance.sendEvent(event)

            if role.defense:
                continue

            isAttack = True

            if role in self.skillTurn:
                if turn in self.skillTurn[role] and role.canSkill():
                    isAttack = False
            else:
                if role.canSkill():
                    isAttack = False

            if isAttack:
                self.doAttack(role, cardList, cardList2)
            else:
                self.doSkill(role, cardList, cardList2)

    # 反击
    def doCounter(self, role: ICard, monster: ICard, cardList2: list[ICard] = []):
        role.doCounter(monster)
        totalDamage = 0
        enemiesBeAttacked = {}
        for record in self.battleListener.findSourceEvent(role, EventType.counter):
            try:
                tempDamage = record.value
                totalDamage += tempDamage
                if record.target in enemiesBeAttacked:
                    tempDamage += enemiesBeAttacked[record.target]
                enemiesBeAttacked[record.target] = tempDamage
            except:
                print('获取反击记录出错')
        role.doBloodSuck(totalDamage)
        for enemy in cardList2:
            if enemy in enemiesBeAttacked:
                if enemiesBeAttacked[enemy] > 0:
                    enemy.beAttacked(enemiesBeAttacked[enemy], False)

    # 必杀
    def doSkill(self, role: ICard, cardList: list[ICard] = [], cardList2: list[ICard] = []):
        role.doSkill()
        totalDamage = 0
        enemiesBeAttacked = {}
        for record in self.battleListener.findSourceEvent(role, EventType.skillDamage):
            try:
                tempDamage = record.value
                totalDamage += tempDamage
                if record.target in enemiesBeAttacked:
                    tempDamage += enemiesBeAttacked[record.target]
                enemiesBeAttacked[record.target] = tempDamage
            except:
                print('获取必杀记录出错')
        for record in self.battleListener.findSourceEvent(role, EventType.skillFollowUp):
            try:
                tempDamage = record.value
                totalDamage += tempDamage
                if record.target in enemiesBeAttacked:
                    tempDamage += enemiesBeAttacked[record.target]
                enemiesBeAttacked[record.target] = tempDamage
            except:
                print('获取必杀追击记录出错')
        role.doBloodSuck(totalDamage)
        for enemy in cardList2:
            if enemy in enemiesBeAttacked:
                if enemiesBeAttacked[enemy] > 0:
                    self.doCounter(enemy, role, cardList)
                    enemy.beAttacked(enemiesBeAttacked[enemy], True)

    # 普攻
    def doAttack(self, role: ICard, cardList: list[ICard] = [], cardList2: list[ICard] = []):
        role.doAttack()
        totalDamage = 0
        enemiesBeAttacked = {}
        for record in self.battleListener.findSourceEvent(role, EventType.attackDamage):
            try:
                tempDamage = record.value
                totalDamage += tempDamage
                if record.target in enemiesBeAttacked:
                    tempDamage += enemiesBeAttacked[record.target]
                enemiesBeAttacked[record.target] = tempDamage
            except:
                print('获取普攻记录出错')
        for record in self.battleListener.findSourceEvent(role, EventType.attackFollowUp):
            try:
                tempDamage = record.value
                totalDamage += tempDamage
                if record.target in enemiesBeAttacked:
                    tempDamage += enemiesBeAttacked[record.target]
                enemiesBeAttacked[record.target] = tempDamage
            except:
                print('获取普攻追击记录出错')
        role.doBloodSuck(totalDamage)
        for enemy in cardList2:
            if enemy in enemiesBeAttacked:
                if enemiesBeAttacked[enemy] > 0:
                    self.doCounter(enemy, role, cardList)
                    enemy.beAttacked(enemiesBeAttacked[enemy], True)

    def writeCardInfoInExcel(self):
        writeCardInfoTitleInExcel(self.roleWS)

        row = 2
        for role in self.team:
            if role.cardName == '临时队友':
                continue
            role.writeCardInfoInExcel(self.roleWS, row)
            row += 1
        return row

    def exportExcel(self, filePath):
        try:
            if self.wb is not None:
                self.wb.save(filePath)
            success = True
        except:
            success = False
        if success:
            self.wb.close()
            self.wb = None
        return success


class BattleRecordListener(EventListener):
    def __init__(self):
        super(BattleRecordListener, self).__init__('战斗数据记录监听器')
        self.battleRecords = []
        self.currentTurn = 0

    def cleanUp(self):
        self.battleRecords = []

    def receiveEvent(self, arg: Event):
        bs = BattleRecord()
        bs.turn = self.currentTurn
        bs.source = arg.data['source']
        bs.valueType = arg.eventType
        bs.value = arg.data['value']
        bs.target = arg.data['target']
        if 'custom' in arg.data:
            bs.custom = arg.data['custom']
        self.battleRecords.append(bs)

    def findSourceEvent(self, role: ICard, et: EventType, ignoreTurn: bool = False):
        if ignoreTurn:
            return [y for y in self.battleRecords if y.source == role and y.valueType == et]
        else:
            return [y for y in self.battleRecords if
                    y.turn == self.currentTurn and y.source == role and y.valueType == et]

    def findTargetEvent(self, role: ICard, et: EventType, ignoreTurn: bool = False):
        if ignoreTurn:
            return [y for y in self.battleRecords if y.target == role and y.valueType == et]
        else:
            return [y for y in self.battleRecords if
                    y.turn == self.currentTurn and y.target == role and y.valueType == et]

    def findCustomEvent(self, role: ICard, et: EventType, ignoreTurn: bool = False):
        if ignoreTurn:
            return [y for y in self.battleRecords if y.custom == role and y.valueType == et]
        else:
            return [y for y in self.battleRecords if
                    y.turn == self.currentTurn and y.custom == role and y.valueType == et]


class BattleRecord:
    def __init__(self):
        self.source: ICard = None
        self.turn = 0
        self.valueType: EventType = None
        self.value = 0
        self.target: ICard = None
        self.custom = None
