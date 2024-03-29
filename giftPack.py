from Nucarnival.costPerformanceHelper import CostPerformanceHelper
from Props.currencyType import CurrencyType
from Props.gameProp import GameProp
from Props.propTypeEnum import PropType, PropTypes
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# 商店礼包
def shopGiftPacks(gps: list[CostPerformanceHelper]):
    xyk = CostPerformanceHelper('小月卡')
    xyk.addGameProp(GameProp(PropType.sorceryGem, 3))
    xyk.addGameProp(GameProp(PropType.basicBoost, 60))
    xyk.currencyType = CurrencyType.eCoin
    xyk.price = 224
    gps.append(xyk)

    dyk = CostPerformanceHelper('大月卡')
    dyk.addGameProp(GameProp(PropType.sorceryGem, 8))
    dyk.addGameProp(GameProp(PropType.spiritGem, 3000))
    dyk.addGameProp(GameProp(PropType.strongBoost, 30))
    dyk.currencyType = CurrencyType.eCoin
    dyk.price = 599
    gps.append(dyk)

    lxyyb = CostPerformanceHelper('旅行应援包')
    lxyyb.addGameProp(GameProp(PropType.spiritGem, 500))
    lxyyb.addGameProp(GameProp(PropType.strongBoost, 10))
    lxyyb.currencyType = CurrencyType.eCoin
    lxyyb.price = 374
    gps.append(lxyyb)

    ctlxyyb = CostPerformanceHelper('长途旅行应援包')
    ctlxyyb.addGameProp(GameProp(PropType.sorceryGem, 6))
    ctlxyyb.addGameProp(GameProp(PropType.basicBoost, 20))
    ctlxyyb.currencyType = CurrencyType.eCoin
    ctlxyyb.price = 374
    gps.append(ctlxyyb)

    xsjqclb = CostPerformanceHelper('新世界启程礼包')
    xsjqclb.addGameProp(GameProp(PropType.essenceContract, 5))
    xsjqclb.addGameProp(GameProp(PropType.coin, 30000))
    xsjqclb.addGameProp(GameProp(PropType.holyWater_S, 30))
    xsjqclb.currencyType = CurrencyType.eCoin
    xsjqclb.price = 299
    gps.append(xsjqclb)

    zfyyb = CostPerformanceHelper('祝福应援礼包')
    zfyyb.addGameProp(GameProp(PropType.spiritGem, 2000))
    zfyyb.addGameProp(GameProp(PropType.coin, 100000))
    zfyyb.addGameProp(GameProp(PropType.holyWater_S, 60))
    zfyyb.addGameProp(GameProp(PropType.holyWater_M, 20))
    zfyyb.currencyType = CurrencyType.eCoin
    zfyyb.price = 674
    gps.append(zfyyb)

    qlkfyylb = CostPerformanceHelper('潜力开发应援礼包')
    qlkfyylb.addGameProp(GameProp(PropType.spiritGem, 1000))
    qlkfyylb.addGameProp(GameProp(PropType.coin, 300000))
    qlkfyylb.addGameProp(GameProp(PropType.basicPotential, 98))
    qlkfyylb.addGameProp(GameProp(PropType.intermediatePotential, 40))
    # 未完成
    qlkfyylb.currencyType = CurrencyType.eCoin
    qlkfyylb.price = 1199
    gps.append(qlkfyylb)

    yslb = CostPerformanceHelper('钥匙礼包（烈焰情深等）')
    yslb.addGameProp(GameProp(PropType.essenceContract, 10))
    yslb.addGameProp(GameProp(PropType.fancyKey, 1))
    yslb.addGameProp(GameProp(PropType.ultraGift, 60))
    yslb.currencyType = CurrencyType.eCoin
    yslb.price = 4224
    gps.append(yslb)

    tmjy = CostPerformanceHelper('甜蜜记忆')
    tmjy.addGameProp(GameProp(PropType.essenceContract, 8))
    tmjy.addGameProp(GameProp(PropType.key, 5))
    tmjy.addGameProp(GameProp(PropType.ultraGift, 30))
    tmjy.currencyType = CurrencyType.eCoin
    tmjy.price = 1675
    gps.append(tmjy)
    return gps


# 主线礼包
def chapterGiftPacks(gps: list[CostPerformanceHelper]):
    c1lb = CostPerformanceHelper('欢庆第1章突破')
    c1lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c1lb.addGameProp(GameProp(PropType.holyWater_M, 10))
    c1lb.currencyType = CurrencyType.eCoin
    c1lb.price = 1049
    gps.append(c1lb)

    c1lbsp = CostPerformanceHelper('欢庆第1章突破SP')
    c1lbsp.addGameProp(GameProp(PropType.spiritGem, 3000))
    c1lbsp.addGameProp(GameProp(PropType.sr_Fragment, 40))
    c1lbsp.addGameProp(GameProp(PropType.holyWater_M, 10))
    c1lbsp.currencyType = CurrencyType.eCoin
    c1lbsp.price = 4224
    gps.append(c1lbsp)

    c2lb = CostPerformanceHelper('欢庆第2章突破')
    c2lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c2lb.addGameProp(GameProp(PropType.sr_Fragment, 24))
    c2lb.addGameProp(GameProp(PropType.holyWater_M, 10))
    c2lb.currencyType = CurrencyType.eCoin
    c2lb.price = 2249
    gps.append(c2lb)

    c2lbsp = CostPerformanceHelper('欢庆第2章突破SP')
    c2lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c2lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c2lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c2lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c2lbsp.currencyType = CurrencyType.eCoin
    c2lbsp.price = 7499
    gps.append(c2lbsp)

    c3lb = CostPerformanceHelper('欢庆第3章突破')
    c3lb.addGameProp(GameProp(PropType.essenceContract, 5))
    c3lb.addGameProp(GameProp(PropType.spiritGem, 150))
    c3lb.addGameProp(GameProp(PropType.holyWater_M, 10))
    c3lb.addGameProp(GameProp(PropType.basicPotential, 5))
    c3lb.currencyType = CurrencyType.eCoin
    c3lb.price = 674
    gps.append(c3lb)

    c3lbsp = CostPerformanceHelper('欢庆第3章突破SP')
    c3lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c3lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c3lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c3lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c3lbsp.currencyType = CurrencyType.eCoin
    c3lbsp.price = 7499
    gps.append(c3lbsp)

    c4lb = CostPerformanceHelper('欢庆第4章突破')
    c4lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c4lb.addGameProp(GameProp(PropType.sr_Fragment, 6))
    c4lb.addGameProp(GameProp(PropType.holyWater_L, 10))
    c4lb.addGameProp(GameProp(PropType.advancedGift, 5))
    c4lb.currencyType = CurrencyType.eCoin
    c4lb.price = 1724
    gps.append(c4lb)

    c4lbsp = CostPerformanceHelper('欢庆第4章突破SP')
    c4lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c4lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c4lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c4lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c4lbsp.currencyType = CurrencyType.eCoin
    c4lbsp.price = 7499
    gps.append(c4lbsp)

    c5lb = CostPerformanceHelper('欢庆第5章突破')
    c5lb.addGameProp(GameProp(PropType.essenceContract, 5))
    c5lb.addGameProp(GameProp(PropType.spiritGem, 300))
    c5lb.currencyType = CurrencyType.eCoin
    c5lb.price = 674
    gps.append(c5lb)

    c5lbsp = CostPerformanceHelper('欢庆第5章突破SP')
    c5lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c5lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c5lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c5lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c5lbsp.currencyType = CurrencyType.eCoin
    c5lbsp.price = 7499
    gps.append(c5lbsp)

    c6lb = CostPerformanceHelper('欢庆第6章突破')
    c6lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c6lb.addGameProp(GameProp(PropType.sr_Fragment, 6))
    c6lb.addGameProp(GameProp(PropType.intermediatePotential, 3))
    c6lb.currencyType = CurrencyType.eCoin
    c6lb.price = 1724
    gps.append(c6lb)

    c6lbsp = CostPerformanceHelper('欢庆第6章突破SP')
    c6lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c6lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c6lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c6lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c6lbsp.currencyType = CurrencyType.eCoin
    c6lbsp.price = 7499
    gps.append(c6lbsp)

    c7lb = CostPerformanceHelper('欢庆第7章突破')
    c7lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c7lb.addGameProp(GameProp(PropType.spiritGem, 1500))
    c7lb.addGameProp(GameProp(PropType.holyWater_M, 10))
    c7lb.addGameProp(GameProp(PropType.intermediatePotential, 5))
    c7lb.currencyType = CurrencyType.eCoin
    c7lb.price = 1724
    gps.append(c7lb)

    c7lbsp = CostPerformanceHelper('欢庆第7章突破SP')
    c7lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c7lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c7lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c7lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c7lbsp.currencyType = CurrencyType.eCoin
    c7lbsp.price = 7499
    gps.append(c7lbsp)

    c8lb = CostPerformanceHelper('欢庆第8章突破')
    c8lb.addGameProp(GameProp(PropType.essenceContract, 5))
    c8lb.addGameProp(GameProp(PropType.holyWater_M, 20))
    c8lb.addGameProp(GameProp(PropType.intermediatePotential, 5))
    c8lb.currencyType = CurrencyType.eCoin
    c8lb.price = 674
    gps.append(c8lb)

    c8lbsp = CostPerformanceHelper('欢庆第8章突破SP')
    c8lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c8lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c8lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c8lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c8lbsp.currencyType = CurrencyType.eCoin
    c8lbsp.price = 7499
    gps.append(c8lbsp)

    c9lb = CostPerformanceHelper('欢庆第9章突破')
    c9lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c9lb.addGameProp(GameProp(PropType.holyWater_L, 10))
    c9lb.addGameProp(GameProp(PropType.advancedPotential, 5))
    c9lb.currencyType = CurrencyType.eCoin
    c9lb.price = 1724
    gps.append(c9lb)

    c9lbsp = CostPerformanceHelper('欢庆第9章突破SP')
    c9lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c9lbsp.addGameProp(GameProp(PropType.ssr_Role, 1))
    c9lbsp.addGameProp(GameProp(PropType.ssr_Fragment, 10))
    c9lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c9lbsp.currencyType = CurrencyType.eCoin
    c9lbsp.price = 7499
    gps.append(c9lbsp)

    c10lb = CostPerformanceHelper('第10章迎战礼包')
    c10lb.addGameProp(GameProp(PropType.spiritGem, 2000))
    c10lb.addGameProp(GameProp(PropType.holyWater_L, 30))
    c10lb.addGameProp(GameProp(PropType.coin, 300000))
    c10lb.addGameProp(GameProp(PropType.strongBoost, 10))
    c10lb.currencyType = CurrencyType.eCoin
    c10lb.price = 1225
    gps.append(c10lb)

    c10lbsp = CostPerformanceHelper('第10章迎战礼包SP')
    c10lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c10lbsp.addGameProp(GameProp(PropType.crystalCore, 1))
    c10lbsp.addGameProp(GameProp(PropType.holyWater_L, 20))
    c10lbsp.currencyType = CurrencyType.eCoin
    c10lbsp.price = 6000
    gps.append(c10lbsp)

    c10lb_2 = CostPerformanceHelper('第10章征战礼包')
    c10lb_2.addGameProp(GameProp(PropType.spiritGem, 2000))
    c10lb_2.addGameProp(GameProp(PropType.holyWater_L, 30))
    c10lb_2.addGameProp(GameProp(PropType.coin, 200000))
    c10lb_2.addGameProp(GameProp(PropType.strongBoost, 20))
    c10lb_2.currencyType = CurrencyType.eCoin
    c10lb_2.price = 1724
    gps.append(c10lb_2)

    c11lb = CostPerformanceHelper('第11章援救礼包')
    c11lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c11lb.addGameProp(GameProp(PropType.coin, 150000))
    c11lb.addGameProp(GameProp(PropType.basicBoost, 10))
    c11lb.currencyType = CurrencyType.eCoin
    c11lb.price = 1724
    gps.append(c11lb)

    c11lbsp = CostPerformanceHelper('第11章援救礼包SP')
    c11lbsp.addGameProp(GameProp(PropType.spiritGem, 3000))
    c11lbsp.addGameProp(GameProp(PropType.essenceContract, 30))
    c11lbsp.addGameProp(GameProp(PropType.crystalCore, 1))
    c11lbsp.currencyType = CurrencyType.eCoin
    c11lbsp.price = 6000
    gps.append(c11lbsp)

    c11lb2 = CostPerformanceHelper('第11章安然礼包')
    c11lb2.addGameProp(GameProp(PropType.essenceContract, 25))
    c11lb2.addGameProp(GameProp(PropType.holyWater_M, 30))
    c11lb2.addGameProp(GameProp(PropType.basicBoost, 20))
    c11lb2.currencyType = CurrencyType.eCoin
    c11lb2.price = 1724
    gps.append(c11lb2)

    c11lbsp2 = CostPerformanceHelper('第11章安然礼包SP')
    c11lbsp2.addGameProp(GameProp(PropType.essenceContract, 40))
    c11lbsp2.addGameProp(GameProp(PropType.coin, 800000))
    c11lbsp2.addGameProp(GameProp(PropType.rawCrystal, 1))
    c11lbsp2.currencyType = CurrencyType.eCoin
    c11lbsp2.price = 6000
    gps.append(c11lbsp2)

    c11lb3 = CostPerformanceHelper('第11章超然礼包')
    c11lb3.addGameProp(GameProp(PropType.ssr_Role, 1))
    c11lb3.addGameProp(GameProp(PropType.coin, 1000000))
    c11lb3.addGameProp(GameProp(PropType.ultraGift, 30))
    c11lb3.currencyType = CurrencyType.eCoin
    c11lb3.price = 6000
    gps.append(c11lb3)

    c12lb = CostPerformanceHelper('第12章释放礼包')
    c12lb.addGameProp(GameProp(PropType.spiritGem, 3000))
    c12lb.addGameProp(GameProp(PropType.essenceContract, 10))
    c12lb.addGameProp(GameProp(PropType.coin, 150000))
    c12lb.currencyType = CurrencyType.eCoin
    c12lb.price = 1245
    gps.append(c12lb)
    
    c12lbsp = CostPerformanceHelper('第12章释放礼包SP')
    c12lbsp.addGameProp(GameProp(PropType.spiritGem, 5000))
    c12lbsp.addGameProp(GameProp(PropType.essenceContract, 40))
    c12lbsp.addGameProp(GameProp(PropType.intermediatePotentialPkg, 10))
    c12lbsp.currencyType = CurrencyType.eCoin
    c12lbsp.price = 6000
    gps.append(c12lbsp)
    
    c12lb2 = CostPerformanceHelper('第12章释然礼包')
    c12lb2.addGameProp(GameProp(PropType.essenceContract, 25))
    c12lb2.addGameProp(GameProp(PropType.coin, 300000))
    c12lb2.addGameProp(GameProp(PropType.intermediatePotentialPkg, 6))
    c12lb2.currencyType = CurrencyType.eCoin
    c12lb2.price = 2979
    gps.append(c12lb2)
    
    c12lb3 = CostPerformanceHelper('第12章探求礼包')
    c12lb3.addGameProp(GameProp(PropType.spiritGem, 2000))
    c12lb3.addGameProp(GameProp(PropType.essenceContract, 15))
    c12lb3.addGameProp(GameProp(PropType.coin, 200000))
    c12lb3.currencyType = CurrencyType.eCoin
    c12lb3.price = 1724
    gps.append(c12lb3)
    
    c12lb3sp = CostPerformanceHelper('第12章探求礼包SP')
    c12lb3sp.addGameProp(GameProp(PropType.spiritGem, 6000))
    c12lb3sp.addGameProp(GameProp(PropType.essenceContract, 40))
    c12lb3sp.addGameProp(GameProp(PropType.intermediatePotentialPkg, 12))
    c12lb3sp.currencyType = CurrencyType.eCoin
    c12lb3sp.price = 6000
    gps.append(c12lb3sp)

    c13lb = CostPerformanceHelper('第13章礼包')
    c13lb.addGameProp(GameProp(PropType.essenceContract, 15))
    c13lb.addGameProp(GameProp(PropType.strongBoost, 10))
    c13lb.addGameProp(GameProp(PropType.holyWater_L, 30))
    c13lb.currencyType = CurrencyType.eCoin
    c13lb.price = 1724
    gps.append(c13lb)

    c13lbsp = CostPerformanceHelper('第13章礼包SP')
    c13lbsp.addGameProp(GameProp(PropType.essenceContract, 40))
    c13lbsp.addGameProp(GameProp(PropType.crystalCore, 1))
    c13lbsp.addGameProp(GameProp(PropType.strongBoost, 30))
    c13lbsp.currencyType = CurrencyType.eCoin
    c13lbsp.price = 6000
    gps.append(c13lbsp)

    c13lb1 = CostPerformanceHelper('第13章祈愿礼包')
    c13lb1.addGameProp(GameProp(PropType.spiritGem, 6000))
    c13lb1.addGameProp(GameProp(PropType.basicBoost, 20))
    c13lb1.addGameProp(GameProp(PropType.intermediatePotentialPkg, 2))
    c13lb1.currencyType = CurrencyType.eCoin
    c13lb1.price = 1245
    gps.append(c13lb1)

    return gps


# 常见活动礼包
def activityGiftPacks(gps: list[CostPerformanceHelper]):
    liandongLB = CostPerformanceHelper('联动礼包')
    liandongLB.addGameProp(GameProp(PropType.essenceContract, 6))
    liandongLB.addGameProp(GameProp(PropType.spiritGem, 2000))
    liandongLB.currencyType = CurrencyType.eCoin
    liandongLB.price = 599
    gps.append(liandongLB)

    clothingLb = CostPerformanceHelper('服装礼包')
    clothingLb.addGameProp(GameProp(PropType.clothing, 1))
    clothingLb.addGameProp(GameProp(PropType.basicBoost, 20))
    clothingLb.addGameProp(GameProp(PropType.crystalCore, 1))
    clothingLb.currencyType = CurrencyType.eCoin
    clothingLb.price = 5120
    gps.append(clothingLb)

    clothingLb2 = CostPerformanceHelper('服装礼包2')
    clothingLb2.addGameProp(GameProp(PropType.clothing, 1))
    clothingLb2.addGameProp(GameProp(PropType.essenceContract, 20))
    clothingLb2.currencyType = CurrencyType.eCoin
    clothingLb2.price = 5120
    gps.append(clothingLb2)

    birthdaylb = CostPerformanceHelper('生日礼包')
    birthdaylb.addGameProp(GameProp(PropType.spiritGem, 1000))
    birthdaylb.addGameProp(GameProp(PropType.bigEssenceVial, 20))
    birthdaylb.addGameProp(GameProp(PropType.ultraGift, 40))
    birthdaylb.currencyType = CurrencyType.eCoin
    birthdaylb.price = 1450
    gps.append(birthdaylb)

    birthdaylb_2 = CostPerformanceHelper('生日语音礼包')
    birthdaylb_2.addGameProp(GameProp(PropType.sorceryGem, 10))
    birthdaylb_2.addGameProp(GameProp(PropType.ultraGift, 30))
    birthdaylb_2.addGameProp(GameProp(PropType.voice, 1))
    birthdaylb_2.currencyType = CurrencyType.eCoin
    birthdaylb_2.price = 1450
    gps.append(birthdaylb_2)

    birthdaylb_3 = CostPerformanceHelper('生日语音礼包2')
    birthdaylb_3.addGameProp(GameProp(PropType.sorceryGem, 20))
    birthdaylb_3.addGameProp(GameProp(PropType.advancedGiftPkg, 3))
    birthdaylb_3.addGameProp(GameProp(PropType.voice, 1))
    birthdaylb_3.currencyType = CurrencyType.eCoin
    birthdaylb_3.price = 1799
    gps.append(birthdaylb_3)

    hdlb1 = CostPerformanceHelper('常见活动礼包1')
    hdlb1.addGameProp(GameProp(PropType.essenceContract, 5))
    hdlb1.addGameProp(GameProp(PropType.coin, 50000))
    hdlb1.currencyType = CurrencyType.eCoin
    hdlb1.price = 345
    gps.append(hdlb1)

    hdlb2 = CostPerformanceHelper('常见活动礼包2')
    hdlb2.addGameProp(GameProp(PropType.essenceContract, 15))
    hdlb2.addGameProp(GameProp(PropType.coin, 100000))
    hdlb2.addGameProp(GameProp(PropType.basicBoost, 20))
    hdlb2.currencyType = CurrencyType.eCoin
    hdlb2.price = 1450
    gps.append(hdlb2)

    hdlb3 = CostPerformanceHelper('常见活动礼包3')
    hdlb3.addGameProp(GameProp(PropType.spiritGem, 1000))
    hdlb3.addGameProp(GameProp(PropType.essenceContract, 30))
    hdlb3.addGameProp(GameProp(PropType.basicBoost, 30))
    hdlb3.currencyType = CurrencyType.eCoin
    hdlb3.price = 2979
    gps.append(hdlb3)

    hdlb4 = CostPerformanceHelper('常见活动礼包4')
    hdlb4.addGameProp(GameProp(PropType.spiritGem, 1000))
    hdlb4.addGameProp(GameProp(PropType.essenceContract, 10))
    hdlb4.addGameProp(GameProp(PropType.coin, 100000))
    hdlb4.currencyType = CurrencyType.eCoin
    hdlb4.price = 1245
    gps.append(hdlb4)

    hdlb5 = CostPerformanceHelper('常见活动礼包5')
    hdlb5.addGameProp(GameProp(PropType.spiritGem, 3000))
    hdlb5.addGameProp(GameProp(PropType.essenceContract, 25))
    hdlb5.addGameProp(GameProp(PropType.coin, 300000))
    hdlb5.currencyType = CurrencyType.eCoin
    hdlb5.price = 2979
    gps.append(hdlb5)

    hdlb6 = CostPerformanceHelper('常见活动礼包6')
    hdlb6.addGameProp(GameProp(PropType.spiritGem, 3500))
    hdlb6.addGameProp(GameProp(PropType.essenceContract, 6))
    hdlb6.currencyType = CurrencyType.eCoin
    hdlb6.price = 1245
    gps.append(hdlb6)

    hdlb7 = CostPerformanceHelper('常见活动礼包7')
    hdlb7.addGameProp(GameProp(PropType.spiritGem, 10000))
    hdlb7.addGameProp(GameProp(PropType.essenceContract, 15))
    hdlb7.addGameProp(GameProp(PropType.intermediatePotentialPkg, 4))
    hdlb7.currencyType = CurrencyType.eCoin
    hdlb7.price = 2979
    gps.append(hdlb7)

    rwlb = CostPerformanceHelper('晶花原石礼包')
    rwlb.addGameProp(GameProp(PropType.essenceContract, 50))
    rwlb.addGameProp(GameProp(PropType.rawCrystal, 1))
    rwlb.currencyType = CurrencyType.eCoin
    rwlb.price = 6920
    gps.append(rwlb)

    bglb = CostPerformanceHelper('背景图礼包')
    bglb.addGameProp(GameProp(PropType.essenceContract, 10))
    bglb.addGameProp(GameProp(PropType.coin, 300000))
    bglb.addGameProp(GameProp(PropType.background, 1))
    bglb.currencyType = CurrencyType.eCoin
    bglb.price = 1660
    gps.append(bglb)

    xclb = CostPerformanceHelper('特殊新春礼包')
    xclb.addGameProp(GameProp(PropType.spiritGem, 2023))
    xclb.addGameProp(GameProp(PropType.essenceContract, 10))
    xclb.currencyType = CurrencyType.eCoin
    xclb.price = 620
    gps.append(xclb)

    fklb1 = CostPerformanceHelper('复刻红钻礼包1')
    fklb1.addGameProp(GameProp(PropType.sorceryGem, 60))
    fklb1.addGameProp(GameProp(PropType.coin, 300000))
    fklb1.addGameProp(GameProp(PropType.basicBoost, 30))
    fklb1.currencyType = CurrencyType.eCoin
    fklb1.price = 2979
    gps.append(fklb1)

    fklb2 = CostPerformanceHelper('复刻红钻礼包2')
    fklb2.addGameProp(GameProp(PropType.sorceryGem, 120))
    fklb2.addGameProp(GameProp(PropType.rawCrystal, 1))
    fklb2.currencyType = CurrencyType.eCoin
    fklb2.price = 6920
    gps.append(fklb2)

    cxlcb = CostPerformanceHelper('初夏里程碑')
    cxlcb.addGameProp(GameProp(PropType.essenceContract, 6))
    cxlcb.addGameProp(GameProp(PropType.advancedGiftPkgForC12, 1))
    cxlcb.currencyType = CurrencyType.eCoin
    cxlcb.price = 599
    gps.append(cxlcb)

    fklb3 = CostPerformanceHelper('复刻礼包1')
    fklb3.addGameProp(GameProp(PropType.essenceContract, 6))
    fklb3.addGameProp(GameProp(PropType.basicBoost, 5))
    fklb3.currencyType = CurrencyType.eCoin
    fklb3.price = 480
    gps.append(fklb3)

    fklb4 = CostPerformanceHelper('复刻礼包2')
    fklb4.addGameProp(GameProp(PropType.essenceContract, 10))
    fklb4.addGameProp(GameProp(PropType.basicBoost, 15))
    fklb4.addGameProp(GameProp(PropType.intermediatePotentialPkg, 2))
    fklb4.currencyType = CurrencyType.eCoin
    fklb4.price = 1245
    gps.append(fklb4)

    znlb = CostPerformanceHelper('新世界畅游（1.5周年礼包）')
    znlb.addGameProp(GameProp(PropType.essenceContract, 3))
    znlb.addGameProp(GameProp(PropType.tinyBoost, 10))
    znlb.currencyType = CurrencyType.eCoin
    znlb.price = 139
    gps.append(znlb)

    jflcb = CostPerformanceHelper('金风里程碑')
    jflcb.addGameProp(GameProp(PropType.spiritGem, 2000))
    jflcb.addGameProp(GameProp(PropType.basicBoost, 15))
    jflcb.addGameProp(GameProp(PropType.coin, 100000))
    jflcb.currencyType = CurrencyType.eCoin
    jflcb.price = 480
    gps.append(jflcb)


# 等级礼包
def lvGiftPacks(gps: list[CostPerformanceHelper]):
    lv15 = CostPerformanceHelper('lv15礼包')
    lv15.addGameProp(GameProp(PropType.spiritGem, 500))
    lv15.addGameProp(GameProp(PropType.coin, 50000))
    lv15.addGameProp(GameProp(PropType.holyWater_L, 30))
    lv15.addGameProp(GameProp(PropType.advancedGift, 3))
    lv15.currencyType = CurrencyType.eCoin
    lv15.price = 224
    gps.append(lv15)

    lv30 = CostPerformanceHelper('lv30礼包')
    lv30.addGameProp(GameProp(PropType.spiritGem, 500))
    lv30.addGameProp(GameProp(PropType.coin, 100000))
    lv30.addGameProp(GameProp(PropType.essenceContract, 3))
    lv30.addGameProp(GameProp(PropType.advancedGift, 5))
    lv30.currencyType = CurrencyType.eCoin
    lv30.price = 299
    gps.append(lv30)

    lv35 = CostPerformanceHelper('lv35礼包')
    lv35.addGameProp(GameProp(PropType.spiritGem, 1000))
    lv35.addGameProp(GameProp(PropType.coin, 120000))
    lv35.addGameProp(GameProp(PropType.key, 1))
    lv35.addGameProp(GameProp(PropType.advancedGift, 8))
    lv35.currencyType = CurrencyType.eCoin
    lv35.price = 674
    gps.append(lv35)

    lv40 = CostPerformanceHelper('lv40礼包')
    lv40.addGameProp(GameProp(PropType.spiritGem, 1000))
    lv40.addGameProp(GameProp(PropType.coin, 120000))
    lv40.addGameProp(GameProp(PropType.essenceContract, 6))
    lv40.currencyType = CurrencyType.eCoin
    lv40.price = 374
    gps.append(lv40)

    lv45 = CostPerformanceHelper('lv45礼包')
    lv45.addGameProp(GameProp(PropType.spiritGem, 1200))
    lv45.addGameProp(GameProp(PropType.coin, 150000))
    lv45.addGameProp(GameProp(PropType.essenceContract, 10))
    lv45.currencyType = CurrencyType.eCoin
    lv45.price = 599
    gps.append(lv45)

    lv50 = CostPerformanceHelper('lv50礼包')
    lv50.addGameProp(GameProp(PropType.spiritGem, 1500))
    lv50.addGameProp(GameProp(PropType.coin, 200000))
    lv50.addGameProp(GameProp(PropType.essenceContract, 10))
    lv50.currencyType = CurrencyType.eCoin
    lv50.price = 599
    gps.append(lv50)

    lv55 = CostPerformanceHelper('lv55礼包')
    lv55.addGameProp(GameProp(PropType.spiritGem, 3000))
    lv55.addGameProp(GameProp(PropType.coin, 250000))
    lv55.addGameProp(GameProp(PropType.rawCrystal, 1))
    lv55.currencyType = CurrencyType.eCoin
    lv55.price = 1725
    gps.append(lv55)

    lv60 = CostPerformanceHelper('lv60礼包')
    lv60.addGameProp(GameProp(PropType.spiritGem, 5000))
    lv60.addGameProp(GameProp(PropType.coin, 300000))
    lv60.addGameProp(GameProp(PropType.rawCrystal, 1))
    lv60.currencyType = CurrencyType.eCoin
    lv60.price = 2475
    gps.append(lv60)


# 魔蕴石
def sorceryGemGiftPacks(gps: list[CostPerformanceHelper]):
    sglb = CostPerformanceHelper('1个魔蕴石')
    sglb.addGameProp(GameProp(PropType.sorceryGem, 1))
    sglb.currencyType = CurrencyType.eCoin
    sglb.price = 74
    gps.append(sglb)

    sglb2 = CostPerformanceHelper('5个魔蕴石')
    sglb2.addGameProp(GameProp(PropType.sorceryGem, 5))
    sglb2.currencyType = CurrencyType.eCoin
    sglb2.price = 369
    gps.append(sglb2)

    sglb3 = CostPerformanceHelper('9个魔蕴石')
    sglb3.addGameProp(GameProp(PropType.sorceryGem, 9))
    sglb3.currencyType = CurrencyType.eCoin
    sglb3.price = 660
    gps.append(sglb3)

    sglb4 = CostPerformanceHelper('20个魔蕴石')
    sglb4.addGameProp(GameProp(PropType.sorceryGem, 20))
    sglb4.currencyType = CurrencyType.eCoin
    sglb4.price = 1465
    gps.append(sglb4)

    sglb5 = CostPerformanceHelper('50个魔蕴石')
    sglb5.addGameProp(GameProp(PropType.sorceryGem, 50))
    sglb5.currencyType = CurrencyType.eCoin
    sglb5.price = 3659
    gps.append(sglb5)

    sglb6 = CostPerformanceHelper('100个魔蕴石')
    sglb6.addGameProp(GameProp(PropType.sorceryGem, 100))
    sglb6.currencyType = CurrencyType.eCoin
    sglb6.price = 7298
    gps.append(sglb6)

    sglb_sp = CostPerformanceHelper('1个魔蕴石（首次）')
    sglb_sp.addGameProp(GameProp(PropType.sorceryGem, 1))
    sglb_sp.addGameProp(GameProp(PropType.spiritGem, 200))
    sglb_sp.currencyType = CurrencyType.eCoin
    sglb_sp.price = 74
    gps.append(sglb_sp)

    sglb2_sp = CostPerformanceHelper('5个魔蕴石（首次）')
    sglb2_sp.addGameProp(GameProp(PropType.sorceryGem, 5))
    sglb2_sp.addGameProp(GameProp(PropType.spiritGem, 1000))
    sglb2_sp.currencyType = CurrencyType.eCoin
    sglb2_sp.price = 369
    gps.append(sglb2_sp)

    sglb3_sp = CostPerformanceHelper('9个魔蕴石（首次）')
    sglb3_sp.addGameProp(GameProp(PropType.sorceryGem, 9))
    sglb3_sp.addGameProp(GameProp(PropType.spiritGem, 1800))
    sglb3_sp.currencyType = CurrencyType.eCoin
    sglb3_sp.price = 660
    gps.append(sglb3_sp)

    sglb4_sp = CostPerformanceHelper('20个魔蕴石（首次）')
    sglb4_sp.addGameProp(GameProp(PropType.sorceryGem, 20))
    sglb4_sp.addGameProp(GameProp(PropType.spiritGem, 4000))
    sglb4_sp.currencyType = CurrencyType.eCoin
    sglb4_sp.price = 1465
    gps.append(sglb4_sp)

    sglb5_sp = CostPerformanceHelper('50个魔蕴石（首次）')
    sglb5_sp.addGameProp(GameProp(PropType.sorceryGem, 50))
    sglb5_sp.addGameProp(GameProp(PropType.spiritGem, 10000))
    sglb5_sp.currencyType = CurrencyType.eCoin
    sglb5_sp.price = 3659
    gps.append(sglb5_sp)

    sglb6_sp = CostPerformanceHelper('100个魔蕴石（首次）')
    sglb6_sp.addGameProp(GameProp(PropType.sorceryGem, 100))
    sglb6_sp.addGameProp(GameProp(PropType.spiritGem, 20000))
    sglb6_sp.currencyType = CurrencyType.eCoin
    sglb6_sp.price = 7298
    gps.append(sglb6_sp)


def initGiftPacks():
    gps: list[CostPerformanceHelper] = []
    shopGiftPacks(gps)
    chapterGiftPacks(gps)
    activityGiftPacks(gps)
    return gps


def exportGiftPacks(ws, title, gps: list[CostPerformanceHelper], needSort=False):
    ws.merge_cells(None, 1, 1, 1, 7)
    ws.cell(1, 1, title)
    row = 2
    ws.cell(row, 1, '礼包')
    ws.cell(row, 2, '物品')
    ws.cell(row, 3, '个数')
    ws.cell(row, 4, '物品等价晶灵石')
    ws.cell(row, 5, '价格（ECoin）')
    ws.cell(row, 6, '基础性价比')
    ws.cell(row, 7, '最优性价比')

    for giftPack in gps:
        giftPack.calCostPerformance()

    if needSort:
        gps.sort(key=lambda x: x.bestCP, reverse=True)

    row += 1
    for giftPack in gps:
        gprow = row
        for prop in giftPack.gamePropList:
            ws.cell(row, 2, prop.propType.typeName)
            ws.cell(row, 3, prop.number)
            result = ''
            if prop in giftPack.basicCPMap:
                result = str(giftPack.basicCPMap[prop].number)
            if prop in giftPack.bestCPMap and prop.propType == PropType.sorceryGem:
                result += '（' + str(giftPack.bestCPMap[prop].number) + '）'
            ws.cell(row, 4, result)
            row += 1
        ws.merge_cells(None, gprow, 1, row - 1, 1)
        ws.cell(gprow, 1, giftPack.giftPackName)

        ws.merge_cells(None, gprow, 5, row - 1, 5)
        ws.cell(gprow, 5, giftPack.price)

        ws.merge_cells(None, gprow, 6, row - 1, 6)
        ws.cell(gprow, 6, str(giftPack.basicCP))
        ws.merge_cells(None, gprow, 7, row - 1, 7)
        result = str(giftPack.bestCP)
        if giftPack.bestCP != giftPack.bestCP2:
            result += '（' + str(giftPack.bestCP2) + '）'
        ws.cell(gprow, 7, result)


def exportAllGiftPacks(filePath):
    wb = Workbook()
    ws = wb.active
    ws.title = '礼包物品性价比计算'
    ws.merge_cells(None, 1, 1, 1, 2)
    ws.merge_cells(None, 2, 1, 2, 2)
    ws.merge_cells(None, 3, 1, 3, 2)
    ws.merge_cells(None, 4, 1, 4, 2)
    ws.merge_cells(None, 5, 1, 5, 2)
    ws.merge_cells(None, 6, 1, 6, 2)
    ws.cell(1, 1, '游戏内各礼包性价比计算')
    ws.cell(2, 1, '所有物品均转为晶灵石后计算（详细请见以下物品转换）')
    ws.cell(3, 1, '部分物品无法直接购买难以定价，例如角色碎片，这些为0晶灵石')
    ws.cell(4, 1, '性价比=礼包物品总价值晶灵石/ecoin')
    ws.cell(5, 1, '基础性价比只考虑抽卡（晶灵石/魔蕴石/魔力契约），其他算为0收益')
    ws.cell(6, 1, '最优性价比考虑所有物品的收益，括号内是魔蕴石转换600晶灵石时的性价比')

    row = 7
    ws.merge_cells(None, row, 1, row, 2)
    ws.cell(row, 1, '物品转换')
    row += 1
    for pt in PropTypes():
        ws.cell(row, 1, pt.typeName)
        ws.cell(row, 2, pt.des)
        row += 1

    ws2 = wb.create_sheet('商店礼包')
    gps: list[CostPerformanceHelper] = []
    shopGiftPacks(gps)
    exportGiftPacks(ws2, '商店礼包', gps)

    ws3 = wb.create_sheet('主线礼包')
    gps: list[CostPerformanceHelper] = []
    chapterGiftPacks(gps)
    exportGiftPacks(ws3, '主线礼包', gps)

    ws4 = wb.create_sheet('等级礼包')
    gps: list[CostPerformanceHelper] = []
    lvGiftPacks(gps)
    exportGiftPacks(ws4, '等级礼包', gps)

    ws5 = wb.create_sheet('活动礼包')
    gps: list[CostPerformanceHelper] = []
    activityGiftPacks(gps)
    exportGiftPacks(ws5, '活动礼包', gps)

    ws6 = wb.create_sheet('魔蕴石购入')
    gps: list[CostPerformanceHelper] = []
    sorceryGemGiftPacks(gps)
    exportGiftPacks(ws6, '魔蕴石购入', gps)

    wb.save(filePath)
    wb.close()


if __name__ == '__main__':
    # exportAllGiftPacks('E:\\新世界\\性价比\\性价比计算.xls')
    exportAllGiftPacks('C:\\fhs\\python\\性价比计算.xls')
