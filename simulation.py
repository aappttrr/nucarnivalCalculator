from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from Common.ncRound import roundDown
from Nucarnival.cardHelper import CardHelper
from Nucarnival.nucarnivalHelper import NucarnivalHelper, roundHalfEven
from RoleCards.buff.buff import Buff
from RoleCards.cards.monster.commonMonster import CommonMonster
from RoleCards.cards.monster.tempTeamMate import TempTeamMate
from RoleCards.cards.olivine.sr_Olivine import SROlivine
from RoleCards.common.card import ICard, writeCardInfoTitleInExcel
from RoleCards.enum.buffTypeEnum import BuffType
from RoleCards.enum.cardOccupationEnum import CardOccupation
from RoleCards.enum.cardRarityEnum import CardRarity
from RoleCards.enum.cardRoleEnum import CardRole
from openpyxl.workbook import Workbook
from RoleCards.enum.passiveEffectivenessDifficultyEnum import PassiveEffectivenessDifficulty


def simulationCombat(filePath, cardHelper: CardHelper, helper: NucarnivalHelper, calGroupRole: bool,
                     forceTeamMate: bool = False, monsterCount: int = 1):
    wb = Workbook()

    column = 15
    ws1 = wb.create_sheet('伤害模拟结果_ssr1星', 0)
    ws2 = wb.create_sheet('伤害模拟结果_ssr3星', 0)
    ws3 = wb.create_sheet('伤害模拟结果_ssr5星', 0)
    ws1.merge_cells(None, 1, 1, 1, column)
    ws2.merge_cells(None, 1, 1, 1, column)
    ws3.merge_cells(None, 1, 1, 1, column)
    title = ''
    if calGroupRole:
        title = '单人13回合期望伤害模拟_群体'
    else:
        title = '单人13回合期望伤害模拟_单体'
    ws1.cell(1, 1, title)
    ws2.cell(1, 1, title)
    ws3.cell(1, 1, title)
    comment = Comment('如果三星被动实战吃满难易程度是中等及以下，则会配置虚拟队友让其吃满被动；否则将吃不满'
                      '\n但类似HP>90%的被动则都会吃满', '纳萨尔')
    ws1.cell(1, 1).comment = comment
    ws2.cell(1, 1).comment = comment
    ws3.cell(1, 1).comment = comment

    row1 = 2
    row2 = 2
    row3 = 2
    exportTitle(ws1, row1)
    exportTitle(ws2, row2)
    exportTitle(ws3, row3)

    if monsterCount > 5:
        monsterCount = 5
    elif monsterCount < 1:
        monsterCount = 1

    for x in cardHelper.cardList:
        if x.rarity == CardRarity.N:
            continue

        # if x.occupation != CardOccupation.Guardian:
        #     continue

        needTeamMate = True
        if x.ped == PassiveEffectivenessDifficulty.difficult or x.ped == PassiveEffectivenessDifficulty.veryDifficult:
            needTeamMate = False
        if forceTeamMate:
            needTeamMate = True

        if x.isAttackGroup == calGroupRole or x.isSkillGroup == calGroupRole:
            row1 += 1
            row2 += 1
            if x.rarity == CardRarity.SSR:
                # 1星6潜
                simulation(helper, needTeamMate, None, ws1, None, x, row1, 1, 5, 6, monsterCount)

                # 3星满潜
                simulation(helper, needTeamMate, None, ws2, None, x, row2, 3, 5, 12, monsterCount)

                row3 += 1
                # 5星满潜
                simulation(helper, needTeamMate, None, ws3, None, x, row3, 5, 5, 12, monsterCount)
            elif x.rarity == CardRarity.SR:
                # 3星6潜
                simulation(helper, needTeamMate, None, ws1, None, x, row1, 3, 5, 6, monsterCount)

                # 5星满潜
                simulation(helper, needTeamMate, None, ws2, None, x, row2, 5, 5, 12, monsterCount)
            else:
                # 3星3潜
                simulation(helper, needTeamMate, None, ws1, None, x, row1, 3, 5, 3, monsterCount)

                # 5星满潜
                simulation(helper, needTeamMate, None, ws2, None, x, row2, 5, 5, 6, monsterCount)

    wb.save(filePath)


def simulationCombat2(filePath, cardHelper: CardHelper, helper: NucarnivalHelper, calGroupRole: bool,
                      forceTeamMate: bool = False, monsterCount: int = 1):
    wb = Workbook()

    column = 15
    ws1 = wb.create_sheet('伤害模拟结果_低练度', 0)
    ws1_ws = wb.create_sheet('伤害模拟结果_低练度_带同练度SR奥', 0)
    ws2 = wb.create_sheet('伤害模拟结果_中练度', 0)
    ws2_ws = wb.create_sheet('伤害模拟结果_中练度_带同练度SR奥', 0)
    ws3 = wb.create_sheet('伤害模拟结果_顶配', 0)
    ws3_ws = wb.create_sheet('伤害模拟结果_顶配_带同练度SR奥', 0)

    ws1.merge_cells(None, 1, 1, 1, column)
    ws2.merge_cells(None, 1, 1, 1, column)
    ws3.merge_cells(None, 1, 1, 1, column)
    ws1_ws.merge_cells(None, 1, 1, 1, column)
    ws2_ws.merge_cells(None, 1, 1, 1, column)
    ws3_ws.merge_cells(None, 1, 1, 1, column)
    title = ''
    title2 = ''
    if calGroupRole:
        title2 = '带同练度SR奥13回合期望伤害模拟_群体'
        title = '单人13回合期望伤害模拟_群体'
    else:
        title2 = '带同练度SR奥13回合期望伤害模拟_单体'
        title = '单人13回合期望伤害模拟_单体'
    ws1.cell(1, 1, title)
    ws2.cell(1, 1, title)
    ws3.cell(1, 1, title)
    ws1_ws.cell(1, 1, title2)
    ws2_ws.cell(1, 1, title2)
    ws3_ws.cell(1, 1, title2)
    comment = Comment('如果三星被动实战吃满难易程度是中等及以下，则会配置虚拟队友让其吃满被动；否则将吃不满'
                      '\n但类似HP>90%的被动则都会吃满', '纳萨尔')
    comment2 = Comment('3CD输出对轴SR奥同时必杀，4CD输出和SR奥同时必杀，6CD输出SR奥对轴在输出前一回合必杀', '纳萨尔')
    ws1.cell(1, 1).comment = comment
    ws2.cell(1, 1).comment = comment
    ws3.cell(1, 1).comment = comment
    ws1_ws.cell(1, 1).comment = comment2
    ws2_ws.cell(1, 1).comment = comment2
    ws3_ws.cell(1, 1).comment = comment2

    row1 = 2
    row2 = 2
    row3 = 2
    exportTitle(ws1, row1)
    exportTitle(ws2, row2)
    exportTitle(ws3, row3)
    exportTitle(ws1_ws, row1)
    exportTitle(ws2_ws, row2)
    exportTitle(ws3_ws, row3)

    srO = SROlivine()
    for x in cardHelper.cardList:
        if x.rarity == CardRarity.N:
            continue

        if x.occupation == CardOccupation.Support or x.occupation == CardOccupation.Guardian \
                or x.occupation == CardOccupation.Healer:
            continue

        needTeamMate = True
        if x.ped == PassiveEffectivenessDifficulty.difficult or x.ped == PassiveEffectivenessDifficulty.veryDifficult:
            needTeamMate = False
        if forceTeamMate:
            needTeamMate = True

        if x.isGroup == calGroupRole:
            row1 += 1
            row2 += 1

            # 低练：SSR1星，SR2星，R3星，蜜话统一4，潜力统一3，等级统一60
            # 中练：SSR2星4房，SR3星5房，R4星5房，潜力统一6，等级统一60
            if x.rarity == CardRarity.SSR:
                # 1星3潜
                srO.setProperties(60, 2, 4, 3)
                simulation(helper, needTeamMate, srO, ws1, ws1_ws, x, row1, 1, 4, 3, monsterCount)

                # 2星4房
                srO.setProperties(60, 3, 5, 6)
                simulation(helper, needTeamMate, srO, ws2, ws2_ws, x, row2, 2, 4, 6, monsterCount)

                row3 += 1
                # 5星满潜
                srO.setProperties(60, 5, 5, 12)
                simulation(helper, needTeamMate, srO, ws3, ws3_ws, x, row3, 5, 5, 12, monsterCount)
            elif x.rarity == CardRarity.SR:
                # 2星3潜
                srO.setProperties(60, 2, 4, 3)
                simulation(helper, needTeamMate, srO, ws1, ws1_ws, x, row1, 2, 4, 3, monsterCount)

                # 3星5房
                srO.setProperties(60, 3, 5, 6)
                simulation(helper, needTeamMate, srO, ws2, ws2_ws, x, row2, 3, 5, 6, monsterCount)
            else:
                # 3星3潜
                srO.setProperties(60, 2, 4, 3)
                simulation(helper, needTeamMate, srO, ws1, ws1_ws, x, row1, 3, 4, 3, monsterCount)

                # 4星5房
                srO.setProperties(60, 3, 5, 6)
                simulation(helper, needTeamMate, srO, ws2, ws2_ws, x, row2, 4, 5, 6, monsterCount)

    wb.save(filePath)


# 进行模拟
def getDamageProportion(helper: NucarnivalHelper, x: ICard, data):
    msg = ''
    attackDamage = data['attackDamage']
    attackFU = data['attackFU']
    skillDamage = data['skillDamage']
    skillFU = data['skillFU']
    dot = data['dot']
    counter = data['counter']
    attackHeal = data['attackHeal']
    skillHeal = data['skillHeal']
    hot = data['hot']
    bloodSuck = data['bloodSuck']
    totalDamage = data['totalDamage']
    totalHeal = data['totalHeal']

    if totalDamage == 0:
        return msg
    if totalDamage > 0:
        attack = attackDamage + attackFU
        skill = skillDamage + skillFU
        if attack > 0:
            proportion = roundHalfEven(attack / totalDamage * 100)
            if len(msg) > 0:
                msg += '\n'
            msg += '普攻({}%)'.format(proportion)
        if skill > 0:
            proportion = roundHalfEven(skill / totalDamage * 100)
            if len(msg) > 0:
                msg += '\n'
            msg += '必杀({}%)'.format(proportion)
        if dot > 0:
            proportion = roundHalfEven(dot / totalDamage * 100)
            if len(msg) > 0:
                msg += '\n'
            msg += '持续伤害({}%)'.format(proportion)
        if counter > 0:
            proportion = roundHalfEven(counter / totalDamage * 100)
            if len(msg) > 0:
                msg += '\n'
            msg += '反击({}%)'.format(proportion)
    return msg


def getRank(x: ICard, damage=0):
    rank = ''
    if x.isAttackGroup or x.isSkillGroup:
        if x.rarity == CardRarity.SSR and x.star == 5:
            if damage >= 150000:
                rank = 'T0'
            elif 100000 <= damage < 150000:
                rank = 'T1'
            else:
                rank = 'T2'
        elif (x.rarity == CardRarity.SSR and x.star == 3) \
                or (x.rarity == CardRarity.SR and x.star == 5) \
                or (x.rarity == CardRarity.R and x.star == 5):
            if damage >= 85000:
                rank = 'T0'
            elif 80000 <= damage < 85000:
                rank = 'T1'
            else:
                rank = 'T2'
        else:
            if damage >= 30000:
                rank = 'T0'
            elif 20000 <= damage < 30000:
                rank = 'T1'
            else:
                rank = 'T2'
    else:
        if x.rarity == CardRarity.SSR and x.star == 5:
            if damage >= 300000:
                rank = 'T0'
            elif 275000 <= damage < 300000:
                rank = 'T1'
            elif 250000 <= damage < 275000:
                rank = 'T2'
            elif 200000 <= damage < 250000:
                rank = 'T3'
            else:
                rank = 'T4'
        elif (x.rarity == CardRarity.SSR and x.star == 3) \
                or (x.rarity == CardRarity.SR and x.star == 5) \
                or (x.rarity == CardRarity.R and x.star == 5):
            if damage >= 200000:
                rank = 'T0'
            elif 175000 <= damage < 200000:
                rank = 'T1'
            elif 150000 <= damage < 175000:
                rank = 'T2'
            elif 100000 <= damage < 150000:
                rank = 'T3'
            else:
                rank = 'T4'
        else:
            if damage >= 100000:
                rank = 'T0'
            elif 80000 <= damage < 100000:
                rank = 'T1'
            elif 70000 <= damage < 80000:
                rank = 'T2'
            elif 65000 <= damage < 70000:
                rank = 'T3'
            else:
                rank = 'T4'
    return rank


def simulation(helper: NucarnivalHelper, needTeamMate: bool, srO: ICard, ws: Worksheet, ws2: Worksheet, x: ICard,
               row, _star, _bond, _tier, monsterCount: int):
    x.setProperties(60, _star, _bond, _tier)
    x.calHpAtk()
    if srO is not None:
        srO.calHpAtk()
    export(ws, x, row)
    if ws2 is not None:
        export(ws2, x, row)

    helper.clearUp()
    if needTeamMate:
        similationTeamMate(helper, x)

    helper.team.append(x)
    for mc in range(0, monsterCount):
        helper.monsters.append(CommonMonster())
    helper.maxTurn = 13
    helper.battleStart(False)
    data = helper.getTotalResult(x)
    column = 14
    ws.cell(row, column, data['totalDamage'])
    column += 1
    ws.cell(row, column, getDamageProportion(helper, x, data))
    # column += 1
    # ws.cell(row, column, getRank(x, data['totalDamage']))

    if srO is not None and ws2 is not None:
        helper.clearUp()
        if needTeamMate:
            similationTeamMate(helper, x)
        helper.team.append(srO)
        if x.skillCD == 3:
            helper.skillTurn[x] = [5, 9, 13]
        elif x.skillCD == 6:
            helper.skillTurn[srO] = [6, 12]
        helper.team.append(x)
        helper.monsters.append(CommonMonster())
        helper.maxTurn = 13
        helper.battleStart(False)
        data = helper.getTotalResult(x)
        column = 14
        ws2.cell(row, column, data['totalDamage'])
        column += 1
        ws2.cell(row, column, getDamageProportion(helper, x, data))
        # column += 1
        # ws2.cell(row, column, getRank(x, data['totalDamage']))


# 导出
def export(ws: Worksheet, x: ICard, row):
    ws.cell(row, 1, x.cardName)
    ws.cell(row, 2, x.nickName)
    ws.cell(row, 3, x.role.roleName)
    ws.cell(row, 4, x.rarity.rarityName)
    ws.cell(row, 5, x.cardType.typeName)
    ws.cell(row, 6, x.occupation.occupationName)
    ws.cell(row, 7, x.lv)
    ws.cell(row, 8, x.star)
    ws.cell(row, 9, x.tier)
    ws.cell(row, 10, x.bond)
    ws.cell(row, 11, x.hp)
    ws.cell(row, 12, x.atk)
    ws.cell(row, 13, x.ped.value)


def exportTitle(ws: Worksheet, row):
    ws.cell(row, 1, '卡')
    ws.cell(row, 2, '昵称')
    ws.cell(row, 3, '角色')
    ws.cell(row, 4, '稀有度')
    ws.cell(row, 5, '类型')
    ws.cell(row, 6, '定位')
    ws.cell(row, 7, '等级')
    ws.cell(row, 8, '星级')
    ws.cell(row, 9, '潜力')
    ws.cell(row, 10, '蜜话')
    ws.cell(row, 11, 'Hp')
    ws.cell(row, 12, 'Atk')
    ws.cell(row, 13, '三星被动实战吃满难易程度')
    ws.cell(row, 14, '13回合单人输出')
    ws.cell(row, 15, '输出占比')
    # ws.cell(row, 16, '梯度')


# 配置模拟队友
def similationTeamMate(helper: NucarnivalHelper, x: ICard):
    if x.cardName == '爵士册封之夜' or x.cardName == '追逐悠远之约' \
            or x.cardName == '守望者的冬季馈礼' or x.cardName == '幽暧新星的祷词':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Striker
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.occupation = CardOccupation.Striker
        helper.team.append(mate2)
    elif x.cardName == '化形宴，飘落叶' or x.cardName == '诡秘妖狐':
        mate = TempTeamMate()
        mate.role = CardRole.Kuya
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Kuya
        helper.team.append(mate2)
    elif x.cardName == '来自主人的新礼物' or x.cardName == '流浪小狼':
        mate = TempTeamMate()
        mate.role = CardRole.Garu
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Garu
        helper.team.append(mate2)
    elif x.cardName == '战斗人偶与回忆' or x.cardName == '魔人偶':
        mate = TempTeamMate()
        mate.role = CardRole.Blade
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Blade
        helper.team.append(mate2)
    elif x.cardName == '白色恋人':
        mate = TempTeamMate()
        mate.role = CardRole.Yakumo
        helper.team.append(mate)
    elif x.cardName == '七叶之花的奇迹':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Support
        helper.team.append(mate)
    elif x.cardName == '华宴夜未央':
        mate = TempTeamMate()
        mate.role = CardRole.Kuya
        helper.team.append(mate)
    elif x.cardName == '偶像实习':
        mate = TempTeamMate()
        mate.role = CardRole.Yakumo
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Olivine
        helper.team.append(mate2)
    elif x.cardName == '永不消散的焰火' or x.cardName == '太阳城主':
        mate = TempTeamMate()
        mate.role = CardRole.Dante
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Dante
        helper.team.append(mate2)
    elif x.cardName == '初露芬芳的甜味':
        mate = TempTeamMate()
        helper.team.append(mate)
    elif x.cardName == '使魔-艾斯特':
        mate = TempTeamMate()
        mate.role = CardRole.Morvay
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Aster
        helper.team.append(mate2)
        mate3 = TempTeamMate()
        mate3.role = CardRole.Aster
        helper.team.append(mate3)
    elif x.cardName == '骑士副团长' or x.cardName == '专属指导':
        mate = TempTeamMate()
        mate.role = CardRole.Edmond
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Edmond
        helper.team.append(mate2)
    elif x.cardName == '古森守护者':
        mate = TempTeamMate()
        mate.role = CardRole.Quincy
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.role = CardRole.Quincy
        helper.team.append(mate2)
    elif x.cardName == '觉醒的晶莹花':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Guardian
        helper.team.append(mate)
    elif x.cardName == '冷艳猩红':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Support
        mate.role = CardRole.Morvay
        helper.team.append(mate)
    elif x.cardName == '狂傲紫魅':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Support
        mate.role = CardRole.Aster
        helper.team.append(mate)
    elif x.cardName == '黔云之血脉':
        mate = TempTeamMate()
        mate.occupation = CardOccupation.Support
        helper.team.append(mate)
        mate2 = TempTeamMate()
        mate2.occupation = CardOccupation.Healer
        helper.team.append(mate2)


def banguaiSimulation(filepath, cardHelper: CardHelper, helper: NucarnivalHelper):
    wb = Workbook()
    ws = wb.active
    ws.title = '角色属性'
    writeCardInfoTitleInExcel(ws)

    # sr昆 必杀工具人 30
    # 暗昆 普攻工具人 19
    # sr墨菲 26

    # 光狼 13
    # 暗团 24
    # sr蛋 34
    # sr狼 32

    # 白八 8
    # 水蛋 15
    # 火团 17
    cList = [30, 19]
    banguaiList = [13, 24, 34, 32, 8, 15, 17]
    skillList = [13, 24, 34, 32, 8, 15]
    attackList = [13, 24, 34, 32, 17]
    row = 2
    for i in cList:
        role = cardHelper.cardList[i]
        if role.rarity == CardRarity.SSR:
            role.setProperties(60, 3, 5, 12)
        else:
            role.setProperties(60, 5, 5, 12)
        role.calHpAtk(True)
        role.writeCardInfoInExcel(ws, row)
        row += 1

    for i in banguaiList:
        role = cardHelper.cardList[i]
        if role.rarity == CardRarity.SSR:
            role.setProperties(60, 3, 5, 12)
        else:
            role.setProperties(60, 5, 5, 12)
        role.calHpAtk(True)
        role.writeCardInfoInExcel(ws, row)
        row += 1

    ws2 = wb.create_sheet('必杀队伍', 1)
    row = 0

    helper.maxTurn = 13
    helper.monsters.append(CommonMonster())

    helper.skillTurn[cardHelper.cardList[32]] = [7, 13]

    helper.team.clear()
    helper.team.append(cardHelper.cardList[30])
    helper.battleStart(False)
    row = saveBattleResult(ws2, helper, row)

    for i in skillList:
        helper.team.clear()
        helper.team.append(cardHelper.cardList[i])
        if i == 24:
            helper.team.append(cardHelper.cardList[26])
        helper.team.append(cardHelper.cardList[30])
        helper.battleStart(False)
        row = saveBattleResult(ws2, helper, row)

    ws3 = wb.create_sheet('普攻队伍', 2)
    row = 0
    helper.skillTurn.clear()
    helper.maxTurn = 14
    helper.skillTurn[cardHelper.cardList[34]] = [5, 9, 13]
    helper.skillTurn[cardHelper.cardList[32]] = [6, 10, 14]

    helper.team.clear()
    helper.team.append(cardHelper.cardList[19])
    helper.battleStart(False)
    row = saveBattleResult(ws3, helper, row)
    for i in attackList:
        helper.team.clear()
        helper.team.append(cardHelper.cardList[i])
        if i == 24:
            helper.team.append(cardHelper.cardList[26])
        helper.team.append(cardHelper.cardList[19])
        helper.battleStart(False)
        row = saveBattleResult(ws3, helper, row)
    wb.save(filepath)


def saveBattleResult(targetWs: Worksheet, helper: NucarnivalHelper, row: int):
    row += 1
    row2 = len(helper.team) + 3
    mt = helper.maxTurn
    for i in range(0, len(helper.team) + 2):
        for column in range(1, mt + 4):
            cellValue = helper.ws.cell(row2, column).value
            if cellValue is not None:
                targetWs.cell(row, column, cellValue)
        row2 += 1
        row += 1
    return row


def tempSimulation(_helper: NucarnivalHelper, _cardHelper: CardHelper):
    for role in _cardHelper.cardList:
        _star = 5
        _tier = 12
        _bond = 5
        if role.rarity == CardRarity.SSR:
            _star = 3
        if role.rarity == CardRarity.R or role.rarity == CardRarity.N:
            _tier = 6
        if role.rarity == CardRarity.N:
            _bond = 0
        role.setProperties(60, _star, _bond, _tier)

    # 辅助
    huaAo = _cardHelper.filterCard('AquaBloom')[0]
    xiaBa = _cardHelper.filterCard('OceanBreeze')[0]
    guaLang = _cardHelper.filterCard('HowlingCyclone')[0]
    shengDan = _cardHelper.filterCard('IcyEquilibrium')[0]
    yiDe = _cardHelper.filterCard('GalacticMist')[0]
    srAo = _cardHelper.filterCard('SROlivine')[0]
    miAo = _cardHelper.filterCard('CaptiveStar')[0]
    shaTuan = _cardHelper.filterCard('FlamingSecret')[0]
    wanAo = _cardHelper.filterCard('CaptiveStar')[0]

    # 输出
    puTuan = _cardHelper.filterCard('KnightlyNight')[0]
    anKun = _cardHelper.filterCard('DistantPromise')[0]
    zaoBa = _cardHelper.filterCard('DarkNova')[0]
    puBa = _cardHelper.filterCard('Homecoming')[0]
    srKun = _cardHelper.filterCard('SRQuincy')[0]
    rKun = _cardHelper.filterCard('RQuincy')[0]
    huoTuan = _cardHelper.filterCard('SweetAroma')[0]
    anTuan = _cardHelper.filterCard('EliteInstructor')[0]
    srAi = _cardHelper.filterCard('SRAster')[0]
    shuiHu = _cardHelper.filterCard('AfternoonDaze')[0]
    huoHu = _cardHelper.filterCard('KitsuneDream')[0]
    xiaHu = _cardHelper.filterCard('AromaticExotica')[0]
    guangLang = _cardHelper.filterCard('EndlessBanquet')[0]
    shuiDan = _cardHelper.filterCard('EternalHanabi')[0]
    xiaDan = _cardHelper.filterCard('ScorchingSun')[0]
    anAo = _cardHelper.filterCard('RadiantAdmiral')[0]
    puKun = _cardHelper.filterCard('AncientCeremony')[0]
    daMo = _cardHelper.filterCard('MauveMayhem')[0]
    daAi = _cardHelper.filterCard('ScarletFinesse')[0]
    srTuan = _cardHelper.filterCard('SREdmond')[0]
    yanBa = _cardHelper.filterCard('ShadowLineage')[0]
    puLian = _cardHelper.filterCard('MidnightOwl')[0]
    puBu = _cardHelper.filterCard('ExplosiveRecall')[0]
    huoBu = _cardHelper.filterCard('CrystalAwakening')[0]

    # 治疗
    puAo = _cardHelper.filterCard('HolyConfession')[0]
    sanKun = _cardHelper.filterCard('BlossomingLegend')[0]
    srBa = _cardHelper.filterCard('SRYakumo')[0]
    yanL = _cardHelper.filterCard('EtherealGuardian')[0]

    # for i in range(1, 15):
    #     if i == 4 or i == 7 or i == 10 or i == 13:
    #         _helper.actionSequence[i] = [4, 1, 3, 2]
    #     else:
    #         _helper.actionSequence[i] = [1, 4, 3, 2]

    # _helper.skillTurn[shaTuan] = [5,8,11]
    # _helper.skillTurn[srAo] = [6, 12]
    # _helper.skillTurn[guaLang] = [6, 12]
    # _helper.skillTurn[wanAo] = [5, 9, 13]
    # _helper.skillTurn[srAo] = [6, 12]
    _helper.skillTurn[shaTuan] = [4,8,11,14,17]
    # _helper.skillTurn[huoBu] = [7,13]

    _helper.maxTurn = 13
    _helper.monsters.append(CommonMonster())
    _helper.team.clear()
    _helper.team.append(yiDe)
    _helper.team.append(yanL)
    # _helper.team.append(shuiHu)
    # _helper.team.append(zaoBa)
    # _helper.team.append(puTuan)

    _helper.battleStart(True)
    name = '必杀队-sr昆-{}.xls'.format(_helper.maxTurn)
    # _helper.exportExcel('E:\\新世界\\攻略\\【2023.9.7】烟岚秘境\\烟八模拟\\' + name)
    # _helper.exportExcel('C:\\fhs\\python\\【2023.9.7】\\' + name)


def starCompareSimulation(_helper: NucarnivalHelper, _cardHelper: CardHelper, turn=13):
    wb = Workbook()
    ws1 = wb.create_sheet('输出', 0)

    name1 = '{}回合单人模拟伤害'.format(turn)
    name2 = '{}回合模拟治疗（5人队取均值）'.format(turn)

    ws1.cell(1, 1, '角色')
    ws1.cell(1, 2, '星级')
    ws1.cell(1, 3, 'Hp')
    ws1.cell(1, 4, 'Atk')
    ws1.cell(1, 5, name1)
    ws1.cell(1, 6, '对比上一星级提升')
    ws1.cell(1, 7, '增攻1152-相当于5星满潜SR奥普攻增攻')
    ws1.cell(1, 8, '增攻提升')
    # ws1.cell(1, 9, '增加造成伤害2%-相当于6潜伊得被动')
    # ws1.cell(1, 10, '增伤提升')
    # ws1.cell(1, 11, '增加必杀伤害17%-相当于3星6潜花奥被动')
    # ws1.cell(1, 12, '增必杀提升')
    # ws1.cell(1, 13, '增加普攻伤害20%-相当于3星圣啖被动')
    # ws1.cell(1, 14, '增普攻提升')

    ws2 = wb.create_sheet('治疗', 1)
    ws2.cell(1, 1, '角色')
    ws2.cell(1, 2, '星级')
    ws2.cell(1, 3, 'Hp')
    ws2.cell(1, 4, 'Atk')
    ws2.cell(1, 5, name2)
    ws2.cell(1, 6, '对比上一星级提升')
    ws2.cell(1, 7, '增攻1152-相当于5星满潜SR奥普攻增攻')
    ws2.cell(1, 8, '增攻提升')
    # ws2.cell(1, 9, '增加造成伤害2%-相当于6潜伊得被动')
    # ws2.cell(1, 10, '增伤提升')
    # ws2.cell(1, 11, '增加必杀伤害17%-相当于3星6潜花奥被动')
    # ws2.cell(1, 12, '增必杀提升')
    # ws2.cell(1, 13, '增加普攻伤害20%-相当于3星圣啖被动')
    # ws2.cell(1, 14, '增普攻提升')

    healRow = 2
    damageRow = 2
    for role in _cardHelper.cardList:
        if role.rarity == CardRarity.N or role.rarity == CardRarity.R:
            continue
        if role.occupation == CardOccupation.Support and role.cardName != '诡夜疾风' and role.cardName != '焰沙暗探的秘梦':
            continue
        if role.occupation == CardOccupation.Healer or role.cardName == '守望者的冬季馈礼':
            healRow = doStarCompareSimulation(role, ws2, _helper, turn, healRow, 1)
            if role.occupation == CardOccupation.Healer:
                continue
        damageRow = doStarCompareSimulation(role, ws1, _helper, turn, damageRow, 0)

    filePath = 'C:\\fhs\\python\\星级对比.xls'
    # filePath = 'E:\\新世界\\战斗模拟\\全角色星级和增攻对比-伤害和治疗.xls'
    wb.save(filePath)


def doStarCompareSimulation(role: ICard, ws: Worksheet, _helper: NucarnivalHelper, turn=13, row=1, calType=0):
    lastRow = row
    ws.merge_cells(None, lastRow, 1, lastRow + 4, 1)
    ws.cell(lastRow, 1, role.nickName)
    lastData = None
    for star in range(1, 6):
        _helper.clearUp()
        role.setProperties(60, star, 5, 12)
        role.calHpAtk()
        ws.cell(lastRow + star - 1, 2, star)
        ws.cell(lastRow + star - 1, 3, role.hp)
        ws.cell(lastRow + star - 1, 4, role.atk)

        similationTeamMate(_helper, role)
        _helper.maxTurn = turn
        _helper.monsters.append(CommonMonster())
        _helper.team.append(role)
        if calType == 0:
            _helper.battleStart(False)
            data = _helper.getTotalResult(role)
            currentData = data['totalDamage']
            ws.cell(lastRow + star - 1, 5, currentData)
            if lastData is not None:
                ws.cell(lastRow + star - 1, 6, currentData / lastData)
            lastData = currentData

            buff = Buff('test1', 1152, 0, BuffType.AtkIncreaseByActualValue)
            buff.isPassive = True
            _helper.battleStart(False, buff)
            data2 = _helper.getTotalResult(role)
            currentData2 = data2['totalDamage']
            ws.cell(lastRow + star - 1, 7, currentData2)
            ws.cell(lastRow + star - 1, 8, currentData2 / currentData)

            # buff = Buff('test2', 0.1, 0, BuffType.DamageIncrease)
            # buff.isPassive = True
            # _helper.battleStart(False, buff)
            # data2 = _helper.getTotalResult(role)
            # currentData2 = data2['totalDamage']
            # ws.cell(lastRow + star - 1, 9, currentData2)
            # ws.cell(lastRow + star - 1, 10, currentData2 / currentData)
            #
            # buff = Buff('test3', 0.1, 0, BuffType.SkillIncrease)
            # buff.isPassive = True
            # _helper.battleStart(False, buff)
            # data2 = _helper.getTotalResult(role)
            # currentData2 = data2['totalDamage']
            # ws.cell(lastRow + star - 1, 11, currentData2)
            # ws.cell(lastRow + star - 1, 12, currentData2 / currentData)
            #
            # buff = Buff('test4', 0.1, 0, BuffType.AttackIncrease)
            # buff.isPassive = True
            # _helper.battleStart(False, buff)
            # data2 = _helper.getTotalResult(role)
            # currentData2 = data2['totalDamage']
            # ws.cell(lastRow + star - 1, 13, currentData2)
            # ws.cell(lastRow + star - 1, 14, currentData2 / currentData)
        elif calType == 1:
            _helper.team.append(TempTeamMate())
            _helper.team.append(TempTeamMate())
            _helper.team.append(TempTeamMate())
            _helper.team.append(TempTeamMate())
            _helper.battleStart(False)
            data = _helper.getTotalResult(role)
            currentData = data['totalHeal']
            currentData = roundDown(currentData / 5)
            ws.cell(lastRow + star - 1, 5, currentData)
            if lastData is not None:
                ws.cell(lastRow + star - 1, 6, currentData / lastData)
            lastData = currentData

            buff = Buff('test1', 1152, 0, BuffType.AtkIncreaseByActualValue)
            buff.isPassive = True
            _helper.battleStart(False, buff)
            data2 = _helper.getTotalResult(role)
            currentData2 = data2['totalHeal']
            currentData2 = roundDown(currentData2 / 5)
            ws.cell(lastRow + star - 1, 7, currentData2)
            ws.cell(lastRow + star - 1, 8, currentData2 / currentData)

            # buff = Buff('test2', 0.1, 0, BuffType.DamageIncrease)
            # buff.isPassive = True
            # _helper.battleStart(False, buff)
            # data2 = _helper.getTotalResult(role)
            # currentData2 = data2['totalHeal']
            # currentData2 = roundDown(currentData2 / 5)
            # ws.cell(lastRow + star - 1, 9, currentData2)
            # ws.cell(lastRow + star - 1, 10, currentData2 / currentData)
            #
            # buff = Buff('test3', 0.1, 0, BuffType.SkillIncrease)
            # buff.isPassive = True
            # _helper.battleStart(False, buff)
            # data2 = _helper.getTotalResult(role)
            # currentData2 = data2['totalHeal']
            # currentData2 = roundDown(currentData2 / 5)
            # ws.cell(lastRow + star - 1, 11, currentData2)
            # ws.cell(lastRow + star - 1, 12, currentData2 / currentData)
            #
            # buff = Buff('test4', 0.1, 0, BuffType.AttackIncrease)
            # buff.isPassive = True
            # _helper.battleStart(False, buff)
            # data2 = _helper.getTotalResult(role)
            # currentData2 = data2['totalHeal']
            # currentData2 = roundDown(currentData2 / 5)
            # ws.cell(lastRow + star - 1, 13, currentData2)
            # ws.cell(lastRow + star - 1, 14, currentData2 / currentData)

    lastRow += 5
    return lastRow


def singleRoleSimulation(_helper: NucarnivalHelper, _cardHelper: CardHelper, roleId, turn=13, lv=60, star=1, bond=5,
                         tier=12):
    role = _cardHelper.filterCard(roleId)[0]
    role.setProperties(lv, star, bond, tier)
    similationTeamMate(_helper, role)
    _helper.maxTurn = turn
    _helper.monsters.append(CommonMonster())
    _helper.team.append(role)
    _helper.battleStart(True)


if __name__ == '__main__':
    _helper = NucarnivalHelper()

    _cardHelper = CardHelper()

    # singleRoleSimulation(_helper, _cardHelper, 'ScarletFinesse', 13, 60, 5)

    tempSimulation(_helper, _cardHelper)

    # starCompareSimulation(_helper, _cardHelper, 13)

    # banguaiSimulation('C:\\fhs\\python\\半拐模拟2.xls', _cardHelper, _helper)

    # simulationCombat('E:\\新世界\\战斗模拟\\单人13回合期望伤害模拟_群体_模拟实战.xls', _cardHelper, _helper, True, False, 1)
    # simulationCombat('E:\\新世界\\战斗模拟\\单人13回合期望伤害模拟_单体_模拟实战.xls', _cardHelper, _helper, False, False, 1)
    # simulationCombat('C:\\fhs\\python\\单人13回合期望伤害模拟_群体_模拟实战2.xls', _cardHelper, _helper, True, False, 1)
    # simulationCombat('C:\\fhs\\python\\单人13回合期望伤害模拟_单体_模拟实战2.xls', _cardHelper, _helper, False, False, 1)
