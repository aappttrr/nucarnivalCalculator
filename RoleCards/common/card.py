from openpyxl.worksheet.worksheet import Worksheet

from Common.cardCalHelper import calBond, calTier, calStar, calLv, calDamageOrHeal
from Common.nvEventManager import eventManagerInstance, EventType, Event
from RoleCards.buff.buff import Buff
from RoleCards.enum.buffTypeEnum import BuffType
from RoleCards.enum.cardOccupationEnum import CardOccupation
from RoleCards.enum.cardRarityEnum import CardRarity
from RoleCards.enum.cardRoleEnum import CardRole
from RoleCards.enum.cardTypeEnum import CardType
from RoleCards.enum.conditionTypeEnum import ConditionType
from RoleCards.enum.passiveEffectivenessDifficultyEnum import PassiveEffectivenessDifficulty
from RoleCards.enum.tierType import TierType
from RoleCards.tier.tierData import getTierData
from Common.ncRound import roundDown, roundCeiling


def writeCardInfoTitleInExcel(ws: Worksheet):
    ws.cell(1, 1, '名称')
    ws.cell(1, 2, '昵称')
    ws.cell(1, 3, '角色')
    ws.cell(1, 4, '属性')
    ws.cell(1, 5, '定位')
    ws.cell(1, 6, 'Hp')
    ws.cell(1, 7, 'Atk')
    ws.cell(1, 8, '等级')
    ws.cell(1, 9, '星级')
    ws.cell(1, 10, '潜能')
    ws.cell(1, 11, '蜜话')


class ICard:
    def __init__(self):
        # id
        self.cardId = None

        # 轮次
        self.round = 1

        # 名称
        self.cardName = None

        # 昵称
        self.nickName = None

        # 简评
        self.des = ''

        # 标签
        self.tag = ''

        # 角色
        self.role: CardRole = None

        # 稀有度
        self.rarity: CardRarity = None

        # 属性
        self.cardType: CardType = CardType.Unset

        # 职业定位
        self.occupation: CardOccupation = None

        # 普通攻击是否为群体攻击
        self.isAttackGroup = False

        # 群体攻击是否为群体攻击
        self.isSkillGroup = False

        # 等级(1-60)
        self.lv = 60

        # 星级
        self.star = 5

        # 生命值
        self.hp = 1

        # 战斗中的当前生命值
        self.hpCurrent = self.hp

        # 最大生命值
        self.maxHp = self.hp

        # 攻击力
        self.atk = 1

        # 蜜话等级(0-5)
        self.bond = 0

        # 潜能(0-12)
        self.tier = 0

        # 潜能类型
        self.tierType: TierType = None

        # 技能CD
        self.skillCD = 0

        # 临时属性，用于计算当前角色是否能释放技能
        self.skillCount = 0

        # 获得的增益
        self.buffs: list[Buff] = []

        # 队友（0-4）
        self.teamMate: list[ICard] = []

        # 敌方（1-5）
        self.enemies: list[ICard] = []

        # lv60 5星 时数据（回廊
        self.lv60s5Hp = 0
        self.lv60s5Atk = 0

        # 当前是否正在防御
        self.defense = False

        # 使用期望数值
        self.useExpectedValue = True

        # 被动吃满难易程度
        self.ped: PassiveEffectivenessDifficulty = None

        # 普攻倍率
        self.attackMagnification = 0
        self.attackHealMagnification = 0

        # 必杀倍率（三个等级）
        self.skillMagnificationLv1 = 0
        self.skillMagnificationLv2 = 0
        self.skillMagnificationLv3 = 0
        self.skillHealMagnificationLv1 = 0
        self.skillHealMagnificationLv2 = 0
        self.skillHealMagnificationLv3 = 0

        self.damageCount = {}
        self.beDamageCount = {}

    def clearUp(self):
        self.skillCount = 0
        self.defense = False
        self.buffs: list[Buff] = []

    def activeBuffs(self):
        self.passive_star_3()
        self.passive_star_5()

    def setProperties(self, _lv, _star, _bond, _tier):
        self.lv = _lv
        self.star = _star
        self.bond = _bond
        self.tier = _tier
        if self.lv <= 0:
            self.lv = 1
        elif self.lv >= 61:
            self.lv = 60
        if self.star <= 0:
            self.star = 1
        elif self.star >= 6:
            self.star = 5
        if self.bond < 0:
            self.bond = 0
        elif self.bond >= 6:
            self.bond = 5

    def setLv(self, _lv):
        self.lv = _lv

    def setStar(self, _star):
        self.star = _star

    def setBond(self, _bond):
        self.bond = _bond

    def setTier(self, _tier):
        self.tier = _tier

    def calHpAtk(self, ignore=False):
        if self.useExpectedValue or ignore:
            self.atk = self.cal(self.lv60s5Atk, True)
            self.hp = self.cal(self.lv60s5Hp, False)

    def calMaxHp(self):
        maxHp = self.hp
        increase = 0
        for buff in self.buffs:
            if buff.buffType == BuffType.HpIncrease:
                increase += buff.value

        if increase != 0:
            maxHp = maxHp * (1 + increase)
            maxHp = roundDown(maxHp)

        self.maxHp = maxHp
        self.hpCurrent = self.maxHp

    def cal(self, lv60s5=1, isAtk=True):
        result = lv60s5

        result = calBond(result, self.bond, self.rarity)

        tierValue = getTierData(self.tierType, self.tier, isAtk)

        result = calTier(result, tierValue)
        result = calStar(result, self.star)
        result = calLv(result, self.lv)
        result = roundCeiling(result)
        return result

    def setHpAtkDirect(self, _hp, _atk):
        self.hp = _hp
        self.atk = _atk

    def setHpDirect(self, _hp):
        self.hp = _hp

    def setAtkDirect(self, _atk):
        self.atk = _atk

    def addDamageCount(self, damage, enemy, isAttack = True):
        if damage > 0:
            self.damageCount[enemy] = 1
            enemy.beDamageCount[self] = 1

    def doAttack(self):
        enemies = self.seizeEnemy(True)
        self.attackBefore(enemies)
        currentAtk = self.getCurrentAtk()
        event1 = Event(EventType.actionAtk)
        event1.data['source'] = self
        event1.data['value'] = currentAtk
        event1.data['target'] = self
        eventManagerInstance.sendEvent(event1)

        self.damageCount = {}
        for tempEnemy in self.enemies:
            self.damageCount[tempEnemy] = 0
            tempEnemy.beDamageCount[self] = 0
        damage = self.attack(enemies, currentAtk)
        if damage > 0:
            if isinstance(enemies, list):
                for enemy in enemies:
                    damage2 = enemy.increaseBeDamage(damage, self, True, False)
                    self.addDamageCount(damage2, enemy, True)
                    event = Event(EventType.attackDamage)
                    event.data['source'] = self
                    event.data['value'] = damage2
                    event.data['target'] = enemy
                    eventManagerInstance.sendEvent(event)
            else:
                damage2 = enemies.increaseBeDamage(damage, self, True, False)
                self.addDamageCount(damage2, enemies, True)
                event = Event(EventType.attackDamage)
                event.data['source'] = self
                event.data['value'] = damage2
                event.data['target'] = enemies
                eventManagerInstance.sendEvent(event)

        heal = self.attackHeal(enemies, currentAtk)
        if heal > 0:
            for mate in self.teamMate:
                heal2 = mate.increaseBeHeal(heal)
                event = Event(EventType.attackHeal)
                event.data['source'] = self
                event.data['value'] = heal2
                event.data['target'] = mate
                eventManagerInstance.sendEvent(event)
                mate.beHealed(heal2, True)

        if damage <= 0 and heal <= 0:
            event = Event(EventType.attack)
            event.data['source'] = self
            event.data['value'] = 0
            event.data['target'] = self
            eventManagerInstance.sendEvent(event)

        self.followUp(currentAtk, True)
        self.triggerWhenAttackOrSkill(enemies, True)

        self.attackAfter(enemies)
        currentAtk2 = self.getCurrentAtk()
        event2 = Event(EventType.actionAfterAtk)
        event2.data['source'] = self
        event2.data['value'] = currentAtk2
        event2.data['target'] = self
        eventManagerInstance.sendEvent(event2)

    def doSkill(self):
        self.skillCount = 0
        enemies = self.seizeEnemy(True)
        self.skillBefore(enemies)
        currentAtk = self.getCurrentAtk()
        event1 = Event(EventType.actionAtk)
        event1.data['source'] = self
        event1.data['value'] = currentAtk
        event1.data['target'] = self
        eventManagerInstance.sendEvent(event1)

        self.damageCount = {}
        for tempEnemy in self.enemies:
            self.damageCount[tempEnemy] = 0
            tempEnemy.beDamageCount[self] = 0
        damage = self.skill(enemies, currentAtk)
        if damage > 0:
            if isinstance(enemies, list):
                for enemy in enemies:
                    damage2 = enemy.increaseBeDamage(damage, self, False, True)
                    self.addDamageCount(damage2, enemy, False)
                    event = Event(EventType.skillDamage)
                    event.data['source'] = self
                    event.data['value'] = damage2
                    event.data['target'] = enemy
                    eventManagerInstance.sendEvent(event)
            else:
                damage2 = enemies.increaseBeDamage(damage, self, False, True)
                self.addDamageCount(damage2, enemies, False)
                event = Event(EventType.skillDamage)
                event.data['source'] = self
                event.data['value'] = damage2
                event.data['target'] = enemies
                eventManagerInstance.sendEvent(event)

        heal = self.skillHeal(enemies, currentAtk)
        if heal > 0:
            for mate in self.teamMate:
                heal2 = mate.increaseBeHeal(heal)
                event = Event(EventType.skillHeal)
                event.data['source'] = self
                event.data['value'] = heal2
                event.data['target'] = mate
                eventManagerInstance.sendEvent(event)
                mate.beHealed(heal2, True)

        if damage <= 0 and heal <= 0:
            event = Event(EventType.skill)
            event.data['source'] = self
            event.data['value'] = 0
            event.data['target'] = self
            eventManagerInstance.sendEvent(event)

        self.followUp(currentAtk, False)
        self.triggerWhenAttackOrSkill(enemies, False)

        self.skillAfter(enemies)
        currentAtk2 = self.getCurrentAtk()
        event2 = Event(EventType.actionAfterAtk)
        event2.data['source'] = self
        event2.data['value'] = currentAtk2
        event2.data['target'] = self
        eventManagerInstance.sendEvent(event2)

    def triggerWhenAttackOrSkill(self, enemies, isAttack: bool):
        for buff in self.buffs:
            if isAttack and buff.conditionType != ConditionType.WhenAttack:
                continue
            elif isAttack is False and buff.conditionType != ConditionType.WhenSkill:
                continue

            if buff.buffType == BuffType.AddDamageIncrease:
                newBuff = Buff(buff.buffId, buff.value, buff.addBuffTurn, BuffType.BeDamageIncrease)
                if isinstance(enemies, list):
                    for enemy in enemies:
                        enemy.addBuff(newBuff, buff.source)
                else:
                    enemies.addBuff(newBuff, buff.source)

    def followUp(self, currentAtk, isAttack: bool):
        for buff in self.buffs:
            if buff.buffType != BuffType.FollowUpAttack:
                continue

            doFollowUp = False
            if isAttack and buff.conditionType == ConditionType.WhenAttack:
                doFollowUp = True
            elif isAttack is False and buff.conditionType == ConditionType.WhenSkill:
                doFollowUp = True
            if doFollowUp is False:
                continue

            if buff.useBaseAtk:
                buffAtk = self.atk
            else:
                buffAtk = currentAtk
            followUpDamage = calDamageOrHeal(buffAtk, buff.value)
            followUpDamage = self.increaseDamage(followUpDamage, buff.seeAsAttack, buff.seeAsSkill)
            enemies = self.seizeEnemy2(buff.isGroup)
            if isinstance(enemies, list):
                for enemy in enemies:
                    damage2 = enemy.increaseBeDamage(followUpDamage, self, buff.seeAsAttack, buff.seeAsSkill)
                    self.addDamageCount(damage2, enemy, isAttack)
                    if isAttack:
                        event = Event(EventType.attackFollowUp)
                    else:
                        event = Event(EventType.skillFollowUp)
                    event.data['source'] = self
                    event.data['value'] = damage2
                    event.data['target'] = enemy
                    eventManagerInstance.sendEvent(event)
            else:
                damage2 = enemies.increaseBeDamage(followUpDamage, self, buff.seeAsAttack, buff.seeAsSkill)
                self.addDamageCount(damage2, enemies, isAttack)
                if isAttack:
                    event = Event(EventType.attackFollowUp)
                else:
                    event = Event(EventType.skillFollowUp)
                event.data['source'] = self
                event.data['value'] = damage2
                event.data['target'] = enemies
                eventManagerInstance.sendEvent(event)

    def getMagnification(self, _s1, _s2, _s4):
        magnification = _s1
        if self.star >= 4:
            magnification = _s4
        elif 2 <= self.star < 4:
            magnification = _s2
        return magnification

    # 反击
    def doCounter(self, targetEnemy):
        currentAtk = self.getCurrentAtk()
        for buff in self.buffs:
            if buff.buffType != BuffType.CounterAttack:
                continue

            if buff.useBaseAtk:
                buffAtk = self.atk
            else:
                buffAtk = currentAtk
            counterDamage = calDamageOrHeal(buffAtk, buff.value)
            counterDamage = self.increaseDamage(counterDamage, buff.seeAsAttack, buff.seeAsSkill)
            enemies = self.seizeEnemy2(buff.isGroup, targetEnemy)
            if isinstance(enemies, list):
                for enemy in enemies:
                    damage2 = enemy.increaseBeDamage(counterDamage, self, buff.seeAsAttack, buff.seeAsSkill)
                    event = Event(EventType.counter)
                    event.data['source'] = self
                    event.data['value'] = damage2
                    event.data['target'] = enemy
                    eventManagerInstance.sendEvent(event)
            else:
                damage2 = enemies.increaseBeDamage(counterDamage, self, buff.seeAsAttack, buff.seeAsSkill)
                event = Event(EventType.counter)
                event.data['source'] = self
                event.data['value'] = damage2
                event.data['target'] = enemies
                eventManagerInstance.sendEvent(event)

    # 吸血
    def doBloodSuck(self, damage):
        if damage > 0:
            totalHeal = 0
            bs = {}
            for buff in self.buffs:
                if buff.buffType == BuffType.BloodSucking:
                    oldBs = 0
                    if buff.source in bs:
                        oldBs = bs.get(buff.source)
                    bs[buff.source] = buff.value + oldBs
            for temp in bs:
                if bs[temp] < 0:
                    continue
                heal = damage * bs[temp]
                heal = roundDown(heal)
                heal = self.increaseBeHeal(heal)
                totalHeal += heal
                if heal > 0:
                    event = Event(EventType.bloodSucking)
                    event.data['source'] = self
                    event.data['value'] = heal
                    event.data['target'] = self
                    event.data['custom'] = temp
                    eventManagerInstance.sendEvent(event)

            if totalHeal > 0:
                self.hpCurrent += totalHeal
                if self.hpCurrent > self.maxHp:
                    self.hpCurrent = self.maxHp

    def sendShieldEvent(self, shield, target):
        event = Event(EventType.shield)
        event.data['source'] = self
        event.data['value'] = shield
        event.data['target'] = target
        eventManagerInstance.sendEvent(event)

    # dot伤害结算：攻击力*倍率*[持续伤害增加(来源于自身，挂上dot的那一刻已确定该伤害)]*[(目标受持续伤害增加+目标受伤害增加)(来源于敌方，在敌方结算dot时计算)]
    # 结算dot，dot是锁面板技能
    def settleDot(self):
        dotDamages = {}
        totalDamage = 0
        for buff in self.buffs:
            if buff.buffType == BuffType.Dot:
                dotDamage = buff.value
                dotDamage = self.increaseBeDot(dotDamage)
                totalDamage += dotDamage
                oldDamage = 0
                if buff.source in dotDamages:
                    oldDamage = dotDamages.get(buff.source)
                dotDamages[buff.source] = dotDamage + oldDamage

        self.beAttacked(totalDamage, False, None)

        for source in dotDamages:
            event = Event(EventType.dot)
            event.data['source'] = source
            event.data['value'] = dotDamages[source]
            event.data['target'] = self
            eventManagerInstance.sendEvent(event)

    # hot治疗结算：攻击力*倍率*[持续治疗增加(来源于治疗的角色)]*[回复量增加(来源于被治疗的角色)]（不一定对）
    # 结算hot，hot是锁面板技能
    def settleHot(self):
        hotHeals = {}
        totalHeal = 0
        for buff in self.buffs:
            if buff.buffType == BuffType.Hot:
                hotHeal = buff.value
                hotHeal = self.increaseBeHot(hotHeal)
                totalHeal += hotHeal
                oldHeal = 0
                if buff.source in hotHeals:
                    oldHeal = hotHeals.get(buff.source)
                hotHeals[buff.source] = hotHeal + oldHeal

        self.beHealed(totalHeal, False)

        for source in hotHeals:
            event = Event(EventType.hot)
            event.data['source'] = source
            event.data['value'] = hotHeals[source]
            event.data['target'] = self
            eventManagerInstance.sendEvent(event)
        return hotHeals

    def getCurrentAtk(self):
        result = self.atk
        increase = 0
        increaseActualValue = 0
        for buff in self.buffs:
            # 如果当前buff条件时，当hp>xx时，当前生命值比例小于条件数值，则不判断该加成
            # 全都不判断，否则会很麻烦
            # 普八、火狐、火团、暗奥、夏狐
            # if buff.conditionType == ConditionType.WhenHpMoreThan:
            #     condition = self.hpCurrent / self.maxHp
            #     if condition <= buff.conditionValue:
            #         continue
            # 获取攻击力增益（百分比）
            if buff.buffType == BuffType.AtkIncrease:
                increase += buff.value
            # 获取攻击力增益（具体数值），一般来源于辅助
            if buff.buffType == BuffType.AtkIncreaseByActualValue:
                increaseActualValue += buff.value

        if increase != 0:
            result = result * (1 + increase)
            result = roundDown(result)
        if increaseActualValue != 0:
            result = result + increaseActualValue

        if result < 0:
            result = 0

        return result

    # 由于dot挂在敌方身上，由敌方结算，所以这里直接获取自身的buff
    def increaseBeDot(self, damage):
        result = damage

        # 由敌方提升的受到持续伤害增加
        dotIncrease = 0

        # 由敌方提升的受到伤害增加
        damageIncrease = 0
        for buff in self.buffs:
            if buff.buffType == BuffType.BeDotIncrease:
                dotIncrease += buff.value

            if buff.buffType == BuffType.BeDamageIncrease:
                damageIncrease += buff.value

        if dotIncrease != 0:
            result = result * (1 + dotIncrease)
            result = roundDown(result)
        if damageIncrease != 0:
            result = result * (1 + damageIncrease)
            result = roundDown(result)

        if result < 0:
            result = 0

        return result

    def increaseDot(self, damage):
        result = damage

        # 由自身提升的持续伤害增加
        dotIncrease = 0

        # 由自身提升的造成伤害增加
        for buff in self.buffs:
            if buff.buffType == BuffType.DotIncrease:
                dotIncrease += buff.value

        if dotIncrease != 0:
            result = result * (1 + dotIncrease)
            result = roundDown(result)

        return result

    def increaseHeal(self, heal, seeAsAttack, seeAsSkill):
        result = heal

        # 由自身提升的造成回复量增加
        healIncrease = 0
        # 由自身提升的普攻/必杀伤害增加
        aIncrease = 0
        sIncrease = 0

        for buff in self.buffs:
            if seeAsAttack and buff.buffType == BuffType.AttackIncrease:
                aIncrease += buff.value
            elif seeAsSkill and buff.buffType == BuffType.SkillIncrease:
                sIncrease += buff.value
            if buff.buffType == BuffType.HealIncrease:
                healIncrease += buff.value

        if aIncrease != 0:
            result = result * (1 + aIncrease)
            result = roundDown(result)
        if sIncrease != 0:
            result = result * (1 + sIncrease)
            result = roundDown(result)
        if healIncrease != 0:
            result = result * (1 + healIncrease)
            result = roundDown(result)
        if result < 0:
            result = 0
        return result

    def increaseBeHeal(self, heal):
        result = heal

        # 由自身提升的持续治疗增加
        healIncrease = 0

        for buff in self.buffs:
            if buff.buffType == BuffType.BeHealIncrease:
                healIncrease += buff.value

        if healIncrease != 0:
            result = result * (1 + healIncrease)
            result = roundDown(result)
        if result < 0:
            result = 0
        return result

    def increaseHot(self, heal):
        result = heal

        # 由自身提升的持续治疗增加
        hotIncrease = 0

        for buff in self.buffs:
            if buff.buffType == BuffType.HotIncrease:
                hotIncrease += buff.value

        if hotIncrease != 0:
            result = result * (1 + hotIncrease)
            result = roundDown(result)

        if result < 0:
            result = 0
        return result

    def increaseBeHot(self, heal):
        result = heal

        # 由自身提升的受持续治疗增加和受回复量增加
        hotIncrease = 0

        for buff in self.buffs:
            if buff.buffType == BuffType.BeHotIncrease:
                hotIncrease += buff.value
            if buff.buffType == BuffType.BeHealIncrease:
                hotIncrease += buff.value

        if hotIncrease != 0:
            result = result * (1 + hotIncrease)
            result = roundDown(result)

        if result < 0:
            result = 0
        return result

    def increaseDamage(self, damage, seeAsAttack, seeAsSkill):
        result = damage
        # 由自身提升的普攻/必杀伤害增加
        aIncrease = 0
        sIncrease = 0

        # 由自身提升的造成伤害增加
        damageIncrease = 0
        for buff in self.buffs:
            if seeAsAttack and buff.buffType == BuffType.AttackIncrease:
                aIncrease += buff.value
            elif seeAsSkill and buff.buffType == BuffType.SkillIncrease:
                sIncrease += buff.value

            if buff.buffType == BuffType.DamageIncrease:
                damageIncrease += buff.value

        if aIncrease != 0:
            result = result * (1 + aIncrease)
            result = roundDown(result)
        if sIncrease != 0:
            result = result * (1 + sIncrease)
            result = roundDown(result)
        if damageIncrease != 0:
            result = result * (1 + damageIncrease)
            result = roundDown(result)

        if result < 0:
            result = 0
        return result

    def increaseBeDamage(self, damage, enemy, seeAsAttack, seeAsSkill):
        result = damage
        # 由自己提升的受到普攻/必杀伤害
        aIncrease = 0
        sIncrease = 0

        # 由自己提升的受到伤害增加
        damageIncrease = 0
        defenseIncrease = 0.5
        damageIncreaseByRole = 0

        for buff in self.buffs:
            if seeAsAttack and buff.buffType == BuffType.BeAttackIncrease:
                aIncrease += buff.value
            elif seeAsSkill and buff.buffType == BuffType.BeSkillIncrease:
                sIncrease += buff.value

            if buff.buffType == BuffType.BeDamageIncrease:
                damageIncrease += buff.value
            if buff.buffType == BuffType.DefenseDamageReduction:
                defenseIncrease += buff.value
            if buff.buffType == BuffType.BeDamageIncreaseByRole \
                    and enemy.role == buff.targetRole:
                damageIncreaseByRole += buff.value

        if aIncrease != 0:
            result = result * (1 + aIncrease)
            result = roundDown(result)
        if sIncrease != 0:
            result = result * (1 + sIncrease)
            result = roundDown(result)
        if damageIncrease != 0:
            result = result * (1 + damageIncrease)
            result = roundDown(result)
        if damageIncreaseByRole != 0:
            result = result * (1 + damageIncreaseByRole)
            result = roundDown(result)
        if self.defense:
            result = result * (1 - defenseIncrease)
            result = roundDown(result)

        # 属性克制
        if enemy.cardType is not None and self.cardType is not None:
            if self.cardType.isBeRestrained(enemy.cardType):
                result = result * 1.2
                result = roundDown(result)
            elif self.cardType.isRestrained(enemy.cardType):
                result = result * 0.8
                result = roundDown(result)

        if result < 0:
            result = 0
        return result

    def increaseShield(self, shield):
        result = shield
        increase = 0
        for buff in self.buffs:
            if buff.buffType == BuffType.ShieldIncrease or buff.buffType == BuffType.ShieldEffectIncrease:
                increase += buff.value

        if increase != 0:
            result = result * (1 + increase)
            result = roundDown(result)

        if result < 0:
            result = 0
        return result

    def increaseBeShield(self, shield):
        result = shield
        increase = 0
        for buff in self.buffs:
            if buff.buffType == BuffType.BeShieldIncrease:
                increase += buff.value

        if increase != 0:
            result = result * (1 + increase)
            result = roundDown(result)

        if result < 0:
            result = 0
        return result

    def isTaunt(self):
        for buff in self.buffs:
            if buff.buffType == BuffType.Taunt:
                return True
        return False

    def disBuffWhenBeAttacked(self):
        tauntBuff: list[Buff] = []
        counterBuff: list[Buff] = []
        disTaunt = False
        disCounter = False
        for buff in self.buffs:
            if buff.buffType == BuffType.Taunt:
                tauntBuff.append(buff)
            elif buff.buffType == BuffType.CounterAttack:
                counterBuff.append(buff)

            if buff.conditionType == ConditionType.WhenBeAttacked:
                if buff.buffType == BuffType.DisTaunt:
                    disTaunt = True
                elif buff.buffType == BuffType.DisCounterAttack:
                    disCounter = True
        if disTaunt:
            for buff in tauntBuff:
                self.buffs.remove(buff)
        if disCounter:
            for buff in counterBuff:
                self.buffs.remove(buff)

    def skillBefore(self, enemies):
        pass

    # 必杀技
    # 必杀伤害结算：攻击力*倍率*[造成伤害增加*必杀伤害增加(来源于自身)]*[目标受必杀伤害增加*目标受伤害增加*属性克制(1.2/1/0.8)(来源于敌方)]
    def skill(self, enemies, currentAtk):
        magnification = self.getMagnification(self.skillMagnificationLv1, self.skillMagnificationLv2,
                                              self.skillMagnificationLv3)
        damage = calDamageOrHeal(currentAtk, magnification)
        damage = self.increaseDamage(damage, False, True)
        return damage

    def skillHeal(self, enemies, currentAtk):
        magnification = self.getMagnification(self.skillHealMagnificationLv1, self.skillHealMagnificationLv2,
                                              self.skillHealMagnificationLv3)
        heal = calDamageOrHeal(currentAtk, magnification)
        heal = self.increaseHeal(heal, False, True)
        return heal

    def skillAfter(self, enemies):
        pass

    def attackBefore(self, enemies):
        pass

    # 普攻
    # 普攻伤害结算：攻击力*倍率*[造成伤害增加*普攻伤害增加(来源于自身)]*[目标受普攻伤害增加*目标受伤害增加*属性克制(1.2/1/0.8)(来源于敌方)]
    def attack(self, enemies, currentAtk):
        damage = calDamageOrHeal(currentAtk, self.attackMagnification)
        damage = self.increaseDamage(damage, True, False)
        return damage

    def attackHeal(self, enemies, currentAtk):
        heal = calDamageOrHeal(currentAtk, self.attackHealMagnification)
        heal = self.increaseHeal(heal, True, False)
        return heal

    def attackAfter(self, enemies):
        pass

    # 3星被动
    def passive_star_3(self):
        if self.star >= 3:
            return True
        return False

    # 5星被动
    def passive_star_5(self):
        if self.star >= 5:
            return True
        return False

    # 信息
    def cardInfo(self, printInfo=True):
        infoStr = self.cardName
        if self.role is not None:
            infoStr = '{} - {}'.format(self.cardName, self.role.roleName)
        if printInfo:
            print(infoStr, end='')
        return infoStr

    def exportCardInfo(self):
        infoStr = self.nickName
        infoStr2 = self.cardName
        if self.role is not None:
            infoStr2 = '{} - {}'.format(self.cardName, self.role.roleName)
        if infoStr is None:
            infoStr = ''
        if infoStr2 is None:
            infoStr2 = ''
        info = infoStr + '\n' + infoStr2
        return info

    def cardDetail(self):
        self.cardInfo()
        print()
        print(
            '等级：{}\n生命值：{}\n攻击力：{}\n蜜话：{}\n潜能：{}'.format(self.lv, self.hp, self.atk, self.bond, self.tier))

    # 是否已阵亡
    def isDead(self):
        if self.hpCurrent <= 0:
            return True
        return False

    # 是否可以释放技能
    def canSkill(self):
        if self.skillCount >= self.skillCD:
            return True
        return False

    def nextRound(self):
        self.skillCount += 1
        self.defense = False
        self.beDamageCount = {}
        removeBuffs: list[Buff] = []
        for buff in self.buffs:
            buff.nextRound()
            if buff.isOver():
                removeBuffs.append(buff)

        for buff in removeBuffs:
            self.buffs.remove(buff)

        removeBuffs.clear()

    def addBuff(self, buff, _source=None):
        if _source is not None:
            buff.source = _source
        else:
            buff.source = self
        self.buffs.append(buff)

    def beAttackedAfter(self, seeAsBeAttacked, source):
        newBuffs = {}
        count = 0
        if source is not None:
            try:
                count = self.beDamageCount[source]
            except:
                count = 0
        if count <= 0:
            count = 1
        if seeAsBeAttacked:
            for buff in self.buffs:
                if buff.conditionType != ConditionType.WhenBeAttacked:
                    continue
                if buff.buffType == BuffType.AddBeDamageIncrease:
                    for i in range(0, count):
                        newBuff = Buff(buff.buffId, buff.value, buff.addBuffTurn, BuffType.BeDamageIncrease)
                        newBuffs[newBuff] = buff.source
        for newBuff in newBuffs:
            self.addBuff(newBuff, newBuffs[newBuff])
        if seeAsBeAttacked:
            self.disBuffWhenBeAttacked()

    def beAttacked(self, damage, seeAsBeAttacked, source):
        if damage <= 0:
            return 0
        removeBuffs: list[Buff] = []
        result = damage
        for buff in self.buffs:
            if buff.buffType == BuffType.Shield:
                if buff.value > result:
                    buff.value -= result
                    result = 0
                elif buff.value == result:
                    removeBuffs.append(buff)
                    result = 0
                else:
                    removeBuffs.append(buff)
                    result -= buff.value

        for buff in removeBuffs:
            self.buffs.remove(buff)

        self.hpCurrent -= result
        self.beAttackedAfter(seeAsBeAttacked, source)
        return 0

    def beHealed(self, heal, seeAsHeal):
        self.hpCurrent += heal
        if self.hpCurrent > self.maxHp:
            self.hpCurrent = self.maxHp

    def calBuffCount(self, _buffId):
        count = 0
        for buff in self.buffs:
            if buff.buffId == _buffId:
                count += 1
        return count

    def writeCardInfoInExcel(self, ws: Worksheet, row: int):
        ws.cell(row, 1, self.cardName)
        ws.cell(row, 2, self.nickName)
        ws.cell(row, 3, self.role.roleName)
        ws.cell(row, 4, self.cardType.typeName)
        ws.cell(row, 5, self.occupation.occupationName)
        ws.cell(row, 6, self.hp)
        ws.cell(row, 7, self.atk)
        ws.cell(row, 8, self.lv)
        ws.cell(row, 9, self.star)
        ws.cell(row, 10, self.tier)
        ws.cell(row, 11, self.bond)

    # 索敌
    def seizeEnemy(self, isAttack: bool, targetEnemy=None):
        isGroup = False
        if isAttack and self.isAttackGroup:
            isGroup = True
        elif isAttack is False and self.isSkillGroup:
            isGroup = True
        return self.seizeEnemy2(isGroup, targetEnemy)

    # 索敌2
    def seizeEnemy2(self, isGroup: bool, targetEnemy=None):
        if isGroup:
            return self.enemies
        else:
            temp1 = [x for x in self.enemies if x.isTaunt()]
            if len(temp1) > 0:
                enemy = temp1[0]
                for temp in temp1:
                    if temp.hpCurrent > enemy.hpCurrent:
                        enemy = temp
                return enemy
            else:
                if targetEnemy is not None:
                    return targetEnemy
            return self.enemies[0]
